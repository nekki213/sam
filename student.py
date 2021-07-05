import os
import pathlib
import re
from datetime import datetime
import pandas as pd
import xlrd
import xlutils.copy
from copy import copy
import openpyxl
import openpyxl.styles
from openpyxl.cell.cell import ILLEGAL_CHARACTERS_RE
from openpyxl.utils.dataframe import dataframe_to_rows
from datetime import datetime


class Student:

    def __init__(self):
        self.key = ''
        self.regno = ''
        self.info = {}


# class ClassList:
#      __init__(sch_year):
#      load_stuinfo()
#      make_dic_print_class_list()
#      make_general_class_list()
#      make_subject_class_list()
#      export_edcity()
#      export_tsa()
#      export_s23()
#      export_sports()
#      export_eca_db()
#      filter_analysis_eca()
#      import_websams()

class ClassList:

    def __init__(self, sch_year: str):
        self.previous_sch_year = str(int(sch_year[0:2])-1) + str(int(sch_year[2:4])-1)
        self.sch_year = sch_year
        self.full_name = sch_year + 'class'
        self.home_folder = os.path.abspath(os.getcwd())
        self.template_folder = os.path.join(self.home_folder, 'template')
        self.classlist_home = os.path.join(self.home_folder, self.full_name)
        self.export_folder = os.path.join(self.classlist_home, 'export')
        self.export_src_folder = os.path.join(self.classlist_home, 'export', 'source')
        self.classlist_file = os.path.join(self.home_folder, self.sch_year + '-classlist.xlsx')
        self.classlist_template_folder = os.path.join(self.home_folder, 'template', 'classlist')

        self.template_file = os.path.join(self.template_folder, 'template.xlsx')
        self.class_template_file = os.path.join(self.classlist_template_folder, 'class-template.xlsx')
        self.print_template_file = os.path.join(self.classlist_template_folder, 'dic-print-template.xlsx')
        self.subject_template_file = os.path.join(self.classlist_template_folder, 'subject-template.xlsx')

        self.template_file_dict = {'class': {'template': self.class_template_file,
                                             'header': ['regno', 'enname', 'chname', 'sex', 'house',
                                                        'classlevel', 'classcode', 'classno',
                                                        'old classlevel', 'old classcode', 'old classno', 'statue']
                                             },
                                   'print': {'template': self.print_template_file,
                                             'header': ['regno', 'enname', 'chname', 'sex', 'house',
                                                        'classlevel', 'classcode', 'classno']
                                             },
                                   'subject': {'template': self.subject_template_file,
                                               'header': ['regno', 'enname', 'chname', 'sex', 'house',
                                                          'classlevel', 'classcode', 'classno'],
                                               'x': ['dh', 'x1', 'x2', 'x3', 'm']
                                               },
                                   }

        self.classlevel = {'s123': ['s1', 's2', 's3'],
                           's456': ['s4', 's5', 's6']}

        self.classcode_dict = {'s1': ['s1a', 's1b', 's1c', 's1d'],
                               's2': ['s2a', 's2b', 's2c', 's2d'],
                               's3': ['s3a', 's3b', 's3c', 's3d'],
                               's4': ['s4a', 's4b', 's4c', 's4d'],
                               's5': ['s5a', 's5b', 's5c', 's5d'],
                               's6': ['s6a', 's6b', 's6c', 's6d']}

        self.stuinfo_df = None

        self.student_sheet_header = ['key', 'regno', 'enname', 'chname', 'sex',
                                     'house', 'classlevel', 'classcode', 'classno', 'old classlevel',
                                     'old classcode', 'old classno', 'statue',
                                     'dh', 'x1', 'x2', 'x3', 'm', 'mth', 'mth_tch']

        self.classlist_header_dict = {'s123': ['regno', 'enname', 'chname', 'sex',
                                               'house', 'classlevel', 'classcode', 'classno', 'old classlevel',
                                               'old classcode', 'old classno', 'statue', 'dh'],
                                      's456': ['regno', 'enname', 'chname', 'sex',
                                               'house', 'classlevel', 'classcode', 'classno', 'old classlevel',
                                               'old classcode', 'old classno', 'statue',
                                               'x1', 'x2', 'x3', 'm'],
                                      }
        self.x_subject = ['m1', 'm2', 'chs', 'hst', 'geo', 'eco', 'phy', 'chm', 'bio', 'ict', 'baf']
        self.x = ['x1', 'x2', 'x3', 'm']

    def load_stuinfo(self):

        print('load stuinfo sheet.')
        self.stuinfo_df = pd.read_excel(self.classlist_file, sheet_name='stuinfo')
        return len(self.stuinfo_df)

    def make_dic_print_class_list(self):

        template_file = self.template_file_dict['print']['template']
        template_header = self.template_file_dict['print']['header']
        template_header_dict = dict(zip(template_header, range(1, len(template_header)+1)))

        print('make dic print class list')
        for forms in ['s123', 's456']:
            for classlevel in self.classlevel[forms]:
                for classcode in self.classcode_dict[classlevel]:

                    save_file = os.path.join(self.classlist_home, 'print', self.sch_year + '-' + classcode + '.xlsx')
                    template_wb = openpyxl.load_workbook(template_file)
                    src_ws = template_wb['src']

                    class_dict = (self.stuinfo_df[(self.stuinfo_df.classcode.isin([classcode.upper()]))]
                                  [template_header]
                                  .applymap(lambda x: re.sub(ILLEGAL_CHARACTERS_RE, '', x) if isinstance(x, str) else x)
                                  .sort_values(by=['classlevel', 'classcode', 'classno'], ascending=[True, True, True])
                                  .to_dict(orient='records')
                                  )
                    row = 2
                    for student in class_dict:
                        for key, item in student.items():
                            src_ws.cell(row=row, column=template_header_dict[key]).value = item
                        else:
                            pass
                            # print('Key({}) not found!'.format(key))
                        row += 1
                    # print(save_file)
                    template_wb.save(save_file)
                    template_wb.close()
                    print('\t{}: {}'.format(classcode, save_file))

    #
    #
    def make_general_class_list(self):

        template_file = self.template_file_dict['class']['template']
        template_header = self.template_file_dict['class']['header']
        # template_header_dict = dict(zip(template_header, range(1, len(template_header) + 1)))

        header_style = openpyxl.styles.NamedStyle(name='header_style')
        header_style.font = openpyxl.styles.Font(bold=True, color='00FFFFFF')
        header_style.fill = openpyxl.styles.PatternFill(fill_type='solid', start_color='000066CC', end_color='000066CC')
        header_style.alignment = openpyxl.styles.Alignment(horizontal='center',
                                                           vertical='center',
                                                           wrap_text=True)

        print('make general class lists')

        classlist_df = (
            self.stuinfo_df[template_header]
                .applymap(lambda x: re.sub(ILLEGAL_CHARACTERS_RE, '', x) if isinstance(x, str) else x)
                .sort_values(by=['classlevel', 'classcode', 'classno'], ascending=[True, True, True])
            )

        classlist_file = os.path.join(self.classlist_home, 'class', self.full_name + '-s123456-all.xlsx')
        classlist_wb = openpyxl.load_workbook(template_file)
        classlist_wb.create_sheet(title='all')

        ws = classlist_wb['all']
        ws.column_dimensions['A'].width = 12
        ws.column_dimensions['B'].width = 20

        for col in ['C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K']:
            ws.column_dimensions[col].width = 8.5

        for row in dataframe_to_rows(classlist_df, index=False, header=True):
            ws.append(row)

        classlist_wb.add_named_style(header_style)
        for cell in ws['1']:
            cell.style = header_style

        for forms in ['s123', 's456']:

            for classlevel in self.classlevel[forms]:

                form_df = (
                    self.stuinfo_df[(self.stuinfo_df.classlevel.isin([classlevel.upper()]))][template_header]
                    .applymap(lambda x: re.sub(ILLEGAL_CHARACTERS_RE, '', x) if isinstance(x, str) else x)
                    .sort_values(by=['classlevel', 'classcode', 'classno'], ascending=[True, True, True])
                )

                save_file = os.path.join(self.classlist_home, 'class', self.full_name + '-' + classlevel + '.xlsx')
                template_wb = openpyxl.load_workbook(template_file)

                template_wb.create_sheet(title=classlevel)
                classlist_wb.create_sheet(title=classlevel)
                classlist_ws = classlist_wb[classlevel]
                ws = template_wb[classlevel]
                ws.column_dimensions['A'].width = 12
                ws.column_dimensions['B'].width = 20

                for col in ['C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K']:
                    ws.column_dimensions[col].width = 8.5

                for row in dataframe_to_rows(form_df, index=False, header=True):
                    ws.append(row)
                    classlist_ws.append(row)

                for cell in ws['1']:
                    cell.style = header_style

                for cell in classlist_ws['1']:
                    cell.style = header_style

                for classcode in self.classcode_dict[classlevel]:

                    template_wb.create_sheet(title=classcode)
                    ws = template_wb[classcode]
                    ws.column_dimensions['A'].width = 12
                    ws.column_dimensions['B'].width = 20

                    for col in ['C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K']:
                        ws.column_dimensions[col].width = 8.5

                    class_df = (
                        self.stuinfo_df[(self.stuinfo_df.classcode.isin([classcode.upper()]))][template_header]
                        .applymap(lambda x: re.sub(ILLEGAL_CHARACTERS_RE, '', x) if isinstance(x, str) else x)
                        .sort_values(by=['classlevel', 'classcode', 'classno'], ascending=[True, True, True])
                        )

                    for row in dataframe_to_rows(class_df, index=False, header=True):
                        ws.append(row)

                    for cell in ws['1']:
                        cell.style = header_style

                template_wb.save(save_file)
                template_wb.close()
                print('\t{}: {}'.format(classlevel, save_file))

        classlist_wb.save(classlist_file)
        classlist_wb.close()
        print('\t{}: {}'.format('all', classlist_file))

    #
    #
    def make_subject_class_list(self):
        date_stamp = datetime.date(datetime.today())
        date_string = '{}.{}.{}'.format(date_stamp.year,
                                        str(date_stamp.month).zfill(2),
                                        str(date_stamp.day).zfill(2))

        template_file = self.template_file_dict['subject']['template']

        template_header = {'s123': self.template_file_dict['subject']['header'] + ['dh'],
                           's456': self.template_file_dict['subject']['header'] + self.x,
                           }

        print('make subject class lists')

        for forms in ['s123', 's456']:

            form_file = os.path.join(self.classlist_home,
                                     'subject',
                                     self.full_name + '-' + forms + '-subject-(' + date_string + ').xlsx')

            form_wb = openpyxl.load_workbook(template_file)

            for classlevel in self.classlevel[forms]:

                form_df = (
                    self.stuinfo_df[(self.stuinfo_df.classlevel.isin([classlevel.upper()]))][template_header[forms]]
                    .applymap(lambda x: re.sub(ILLEGAL_CHARACTERS_RE, '', x) if isinstance(x, str) else x)
                    .sort_values(by=['classlevel', 'classcode', 'classno'], ascending=[True, True, True])
                    )

                form_wb.create_sheet(title=classlevel)
                form_ws = form_wb[classlevel]

                for row in dataframe_to_rows(form_df, index=False, header=True):
                    form_ws.append(row)

                if forms == 's456':

                    subject_file = os.path.join(self.classlist_home,
                                                'subject',
                                                self.full_name + '-' + classlevel
                                                + '-subject-(' + date_string + ').xlsx')
                    subject_wb = openpyxl.load_workbook(template_file)

                    for x_ in self.x:

                        x_header = self.template_file_dict['subject']['header'] + [x_]

                        for subject in self.x_subject:

                            x_subject = x_ + subject
                            x_subject_df = (
                                form_df[(form_df[x_] == subject)][x_header]
                                .sort_values(by=['classlevel', 'classcode', 'classno'], ascending=[True, True, True])
                                )

                            if not x_subject_df.empty:

                                subject_wb.create_sheet(title=x_subject)
                                x_subject_ws = subject_wb[x_subject]

                                for row in dataframe_to_rows(x_subject_df, index=False, header=True):
                                    x_subject_ws.append(row)

                            else:
                                pass

                    subject_wb.save(subject_file)
                    subject_wb.close()
                    print('\t{}: {}'.format('all', subject_file))

            if forms == 's456':

                classlist_wb_src = openpyxl.load_workbook(filename=self.classlist_file, data_only=True)
                sheet_list = classlist_wb_src.sheetnames

                if 'xstat' in sheet_list:

                    ws_src = classlist_wb_src['xstat']
                    # ws_new = wb_new.active
                    ws_new = form_wb.create_sheet(title='stat')
                    # copy worksheet attributes
                    ws_new.sheet_format = copy(ws_src.sheet_format)
                    ws_new.sheet_format = copy(ws_src.sheet_format)
                    ws_new.sheet_properties = copy(ws_src.sheet_properties)
                    ws_new.merged_cells = copy(ws_src.merged_cells)
                    ws_new.page_margins = copy(ws_src.page_margins)
                    ws_new.page_setup = copy(ws_src.page_setup)
                    ws_new.print_options = copy(ws_src.print_options)
                    ws_new.row_dimensions = copy(ws_src.row_dimensions)
                    ws_new.column_dimensions = copy(ws_src.column_dimensions)
                    ws_new._print_area = copy(ws_src._print_area)

                    # copy cell by cell
                    for row in ws_src.rows:

                        for cell in row:

                            new_cell = ws_new.cell(row=cell.row, column=cell.col_idx, value=cell.value)

                            if cell.has_style:

                                new_cell.font = copy(cell.font)
                                new_cell.border = copy(cell.border)
                                new_cell.fill = copy(cell.fill)
                                new_cell.number_format = copy(cell.number_format)
                                new_cell.protection = copy(cell.protection)
                                new_cell.alignment = copy(cell.alignment)

            form_wb.save(form_file)
            form_wb.close()
            print('\t{}: {}'.format('all subject', form_file))

    def export_edcity(self):

        import_header = ['學生註冊編號', '英文姓名', '中文姓名', '顯示姓名 (由系統建立)', '性別', '級別', '班別', '班號',
                         '學校電郵', '國家地區代碼', '流動電話', '登入帳號', '預設密碼']

        import_file = os.path.join(self.classlist_home, 'export', 'edcity', 'studentlist_20200902091212_5.xlsx')
        import_df = pd.read_excel(import_file, '學生名單')[['學生註冊編號', '顯示姓名 (由系統建立)']]
        import_df.rename(columns={'學生註冊編號': 'regno'}, inplace=True)

        filter_column = ['regno', 'enname', 'chname', 'sex', 'classlevel', 'classcode', 'classno', 'pw']
        edcity_df = (
            self.stuinfo_df[filter_column]
                .applymap(lambda x: re.sub(ILLEGAL_CHARACTERS_RE, '', x) if isinstance(x, str) else x)
                .sort_values(by=['classlevel', 'classcode', 'classno'], ascending=[True, True, True])
            )

        edcity_df = pd.merge(edcity_df, import_df, how='left', on='regno')
        edcity_df['classcode'] = edcity_df['classcode'].apply(lambda x: x[1:].upper())
        edcity_df['email'] = edcity_df['regno'].apply(lambda x: x.lower() + '@school.cdgfss.edu.hk')
        edcity_df['account'] = edcity_df['regno'].apply(lambda x: x[0:3].lower() + '-' + x[3:])
        edcity_df['學校電郵'] = ''
        edcity_df['國家地區代碼'] = ''
        edcity_df['流動電話'] = ''

        # edcity_df['regno'] = edcity_df['regno']

        # for various exporting
        edcity_header_c = ['學生註冊編號', '英文姓名', '中文姓名', '性別', '級別', '班別',
                           '班號', '登入帳號', '預設密碼']
        edcity_header_e = ['regno', 'enname', 'chname', 'sex', 'classlevel', 'classcode',
                           'classno', 'account', 'pw']
        edcity_dict = dict(zip(edcity_header_e, edcity_header_c))
        # edcity:   #CDG140126   CHENG HOI YAN   鄭鎧忻   S1   1A   1   cdg-140126   296478

        export_header = ['學生註冊編號', '英文姓名', '中文姓名',  '顯示姓名 (由系統建立)', '性別', '級別', '班別',
                         '班號', '學校電郵', '國家地區代碼', '流動電話', '登入帳號', '預設密碼']
        save_file = os.path.join(self.classlist_home, 'export', 'edcity', self.full_name + '-edcity.xlsx')
        (edcity_df.rename(columns=edcity_dict)
                  .reindex(columns=export_header)
                  .to_excel(save_file, index=False)
         )
        return 0

    def export_tsa(self):

        # for various exporting
        print('******************************************************')
        print('*        need to update s3 withdrawn students        *')
        print('*        in stuinfo sheet before generating          *')
        print('*        tsa list                                    *')
        print('******************************************************')

        print('\ngenerate tsa list')
        filter_column = ['enname', 'chname', 'sex', 'classlevel', 'classcode', 'classno', 'dob', 'strn']
        # School Year,  CLASSLEVEL, CLASSCODE, CLASSNO, ENNAME, CHNAME, SEX, DOB, STRN
        # 2019/2020, S3, 3A, 1, xxx, xxx, M, dd/mm/yyyy, Axxxxxxx
        tsa_column = ['schyear', 'classlevel', 'classcode', 'classno', 'enname', 'chname', 'sex', 'dob2', 'strn']

        tsa_df = (self.stuinfo_df[self.stuinfo_df.classlevel.isin(['S3'])][filter_column]
                      .applymap(lambda x: re.sub(ILLEGAL_CHARACTERS_RE, '', x) if isinstance(x, str) else x)
                      .sort_values(by=['classlevel', 'classcode', 'classno'], ascending=[True, True, True])
                  )

        tsa_df['schyear'] = '20' + self.sch_year[:2] + '/20' + self.sch_year[2:4]
        tsa_df['classcode'] = tsa_df['classcode'].apply(lambda x: x[1:].upper())

        tsa_df['year'] = tsa_df['dob'].str.split('-').str[0]
        tsa_df['month'] = tsa_df['dob'].str.split('-').str[1].str.zfill(2)
        tsa_df['day'] = tsa_df['dob'].str.split('-').str[2].str.zfill(2)

        tsa_df['dob2'] = tsa_df['day'] + '/' + tsa_df['month'] + '/' + tsa_df['year']

        save_file_csv = os.path.join(self.classlist_home, 'export', 'tsa', self.full_name + '-s3-tsa.csv')
        save_file_xls = os.path.join(self.classlist_home, 'export', 'tsa', self.full_name + '-s3-tsa.xls')

        tsa_df[tsa_column].to_csv(save_file_csv, sep='\t', header=False, index=False, encoding='utf-8')
        tsa_df[tsa_column].to_excel(save_file_xls, index=False)
        print('\ttsa file: {} / {}'.format(save_file_csv, save_file_xls))

        return 0

    def export_s23_final(self):

        filter_column = ['regno', 'enname', 'chname', 'sex', 'house',
                         'classlevel', 'classcode', 'classno',
                         'old classlevel', 'old classcode', 'old classno']
        result_column = ['regno', 'eng', 'mth']

        student_df = (self.stuinfo_df[self.stuinfo_df.classlevel.isin(['S2', 'S3'])][filter_column]
                      .applymap(lambda x: re.sub(ILLEGAL_CHARACTERS_RE, '', x) if isinstance(x, str) else x)
                      .sort_values(by=['classlevel', 'classcode', 'classno'], ascending=[True, True, True])
                      )
        result_df = pd.read_excel(self.classlist_file, sheet_name='s123final')[result_column]

        stu_result_df = pd.merge(student_df, result_df, how='left', on='regno')
        #  print(stu_result_df.head(5))

        for subject in ['eng', 'mth']:
            for classlevel in ['S2', 'S3']:
                save_file = os.path.join(self.classlist_home, 'export',
                                         subject,
                                         self.full_name + '-' + classlevel.lower() + '-' + subject + '.xlsx')
                (stu_result_df[stu_result_df['classlevel'] == classlevel][filter_column + [subject]]
                 .to_excel(save_file, index=False)
                 )
        return 0

    def export_sports(self):

        # for various exporting
        # 'classlevel', 'classcode', 'classno', 'enname', 'chname', 'dob', 'house', 'division', 'sex', 'hkid', 'noid',
        print('\ngenerate sports list')
        filter_column = ['enname', 'chname', 'sex', 'classlevel', 'classcode', 'classno', 'house', 'dob', 'hkid']

        # S1, S1A, 1, CHAN MAN CHING, 陳敏晴, 28/03/2007, J, D, F, S1413289,	No-ID
        sports_column = ['classlevel', 'classcode', 'classno', 'enname', 'chname',
                         'dob2', 'house', 'division', 'sex', 'hkid', 'noid']

        sports_df = (self.stuinfo_df[filter_column]
                     .applymap(lambda x: re.sub(ILLEGAL_CHARACTERS_RE, '', x) if isinstance(x, str) else x)
                     .sort_values(by=['classlevel', 'classcode', 'classno'], ascending=[True, True, True])
                     )

        sports_df['year'] = sports_df['dob'].str.split('-').str[0]
        sports_df['month'] = sports_df['dob'].str.split('-').str[1].str.zfill(2)
        sports_df['day'] = sports_df['dob'].str.split('-').str[2].str.zfill(2)
        sports_df['dob2'] = sports_df['day'] + '/' + sports_df['month'] + '/' + sports_df['year']
        sports_df['division'] = 'D'
        sports_df['noid'] = 'NO-ID'

        save_file_csv = os.path.join(self.classlist_home, 'export', 'sports', self.full_name + '-sports.csv')
        sports_df[sports_column].to_csv(save_file_csv, header=False, index=False, encoding='utf_8_sig')

        print('\tsports file: {}'.format(save_file_csv))
        print('\tNeed to re-save in non utf-8 encoding by excel/notepad.')
        return 0

    def export_eca_db(self):
        # 1. build new student records
        # 2. build new student accounts
        # 3. promote old student classes

        sch_year = '20' + self.sch_year[:2]
        print('generate eac database update sql')

        filter_header = ['regno', 'enname', 'chname', 'sex', 'house',
                         'classlevel', 'classcode', 'classno',
                         'old classlevel', 'old classcode', 'pw']

        student_df = (
            self.stuinfo_df[filter_header]
                .applymap(lambda x: re.sub(ILLEGAL_CHARACTERS_RE, '', x) if isinstance(x, str) else x)
                .sort_values(by=['classlevel', 'classcode', 'classno'], ascending=[True, True, True])
            )

        student_df['old classcode'].fillna('-1', inplace=True)
        new_student_df = student_df[student_df['old classcode'] == '-1'].sort_values(by=['regno'], ascending=[True])
        # print(len(student_df))
        # print(len(new_student_df))
        new_student_df['email'] = new_student_df['regno'].apply(lambda x: str(x).lower() + '@school.cdgfss.edu.hk')
        password = '52c69e3a57331081823331c4e69d3f2e'
        new_student_dict = new_student_df.to_dict(orient='records')

        eac_sql1 = os.path.join(self.classlist_home, 'export', 'eac', self.full_name + '_eac_new_users_sql.txt')
        eac_sql2 = os.path.join(self.classlist_home, 'export', 'eac', self.full_name + '_eac_new_students_sql.txt')
        eac_sql3 = os.path.join(self.classlist_home, 'export', 'eac', self.full_name + '_eac_update_studentclass_sql.txt')
        eac_sql4 = os.path.join(self.classlist_home, 'export', 'eac', self.full_name + '_eac_new_teacher_sql.txt')
        eac_sql5 = os.path.join(self.classlist_home, 'export', 'eac', self.full_name + '_eac_new_teacher_user_sql.txt')
        eac_pw_csv = os.path.join(self.classlist_home, 'export', 'eac', self.full_name + '_eac_pw.csv')

        sql_out1 = open(eac_sql1, 'w', encoding='utf-8')
        sql_out2 = open(eac_sql2, 'w', encoding='utf-8')
        sql_out3 = open(eac_sql3, 'w', encoding='utf-8')
        sql_out4 = open(eac_sql4, 'w', encoding='utf-8')
        sql_out5 = open(eac_sql5, 'w', encoding='utf-8')


        for student in new_student_dict:
            if student['regno'] is None:
                pass
            else:
                # print(student['regno'])
                sql_txt1 = ('insert into tbl_users(userid, password, email, type) '
                            'values (\'{}\', \'{}\', \'{}\', \'{}\');\n'
                            .format(student['regno'], password, student['email'], 'student'))

                sql_txt2 = ('insert into tbl_students(studentid, ename, cname, gender, house, remedial, status) '
                            'values (\'{}\', \'{}\', \'{}\', \'{}\', \'{}\', \'{}\', \'{}\');\n'
                            .format(student['regno'], student['enname'], student['chname'],
                                    student['sex'], student['house'], 'no', 'active'))

                sql_out1.write(sql_txt1)
                sql_out2.write(sql_txt2)

        print('\tnew user sql: {}'.format(eac_sql1))
        print('\tnew new student sql: {}'.format(eac_sql2))

        sql_out1.close()
        sql_out2.close()

        for student in student_df.to_dict(orient='records'):
            if student['regno'] is None:
                pass
            else:
                # print(student['regno'])
                sql_txt3 = ('insert into tbl_studentclasses(studentid, year, class, no) '
                            'values (\'{}\', \'{}\', \'{}\', \'{}\');\n'
                            .format(student['regno'], sch_year, student['classcode'], student['classno']))

                sql_out3.write(sql_txt3)

        print('\tnew class sql: {}'.format(eac_sql3))
        sql_out3.close()
        x = 'x'
        teacher_dict = [{'tch_code': '20lj', 'teacherid': 'LJ', 'cname': x, 'ename': 'Leung Ka Shing', 'gender': 'M'},
                        {'tch_code': '20lb', 'teacherid': 'LB', 'cname': x, 'ename': 'Leung Ming Kan', 'gender': 'M'},
                        {'tch_code': '20hl', 'teacherid': 'HL', 'cname': x, 'ename': 'Lim Wing Hin Henry', 'gender': 'M'},
                        {'tch_code': '20km', 'teacherid': 'KM', 'cname': x, 'ename': 'Mak Hoi Ying', 'gender': 'M'},
                        {'tch_code': '20as', 'teacherid': 'AS', 'cname': x, 'ename': 'So Chun Kit Ambrose', 'gender': 'M'},
                        ]

        for teacher in teacher_dict:
            sql_txt4 = ('insert into tbl_teachers(teacherid, cname, ename, gender) '
                        'values (\'{}\', \'{}\', \'{}\', \'{}\');\n'
                        .format(teacher['teacherid'], teacher['cname'], teacher['ename'], teacher['gender']))

            sql_out4.write(sql_txt4)

        print('\tnew teacher sql: {}'.format(eac_sql4))
        sql_out4.close()


        for teacher in teacher_dict:
            email_txt = 't' + teacher['tch_code'] + '@school.cdgfss.edu.hk'
            sql_txt5 = ('insert into tbl_users(userid, password, email, type) '
                        'values (\'{}\', \'{}\', \'{}\', \'{}\');\n'
                        .format(teacher['tch_code'], password, email_txt, 'teacher'))
            sql_out5.write(sql_txt5)

        print('\tnew teacher user sql: {}'.format(eac_sql5))
        sql_out5.close()

        student_df[['regno', 'pw']].to_csv(eac_pw_csv)
        print('\tpassword csv: {}'.format(eac_pw_csv))

        return 0

    def export_jupas_eapp(self):

        print('\ngenerate S6 jupas list')
        filter_column = ['key', 'regno', 'enname', 'classlevel', 'classcode', 'classno', 'hkid']
        jupas_column = ['HKID Card No.', 'School Code',
                        'Last Name', 'First Name', 'Class',
                        'Student No. (Optional)', 'Group (Optional)',
                        'Passport No. (Optional)', 'Passport Issuing Country (Optional)']

        # School Code	HKID	Passport Number	Issue Country	Last Name	First Name	Class	Student Number	Group
        # 20212	        Y768264(4)			                    WONG	    Yee Kei	    S6C	    CDG	            S6C27
        # should use regno for 'Student Number' instead of the class no.

        jupas_df = (self.stuinfo_df[self.stuinfo_df['classlevel'] == 'S6'][filter_column]
                    .applymap(lambda x: re.sub(ILLEGAL_CHARACTERS_RE, '', x) if isinstance(x, str) else x)
                    .sort_values(by=['classlevel', 'classcode', 'classno'], ascending=[True, True, True])
                    )

        eapp_df = (self.stuinfo_df[self.stuinfo_df['classlevel'] == 'S6'][filter_column]
                    .applymap(lambda x: re.sub(ILLEGAL_CHARACTERS_RE, '', x) if isinstance(x, str) else x)
                    .sort_values(by=['classlevel', 'classcode', 'classno'], ascending=[True, True, True])
                    )

        jupas_df['HKID Card No.'] = jupas_df['hkid'].str[0:7] + '(' + jupas_df['hkid'].str[7] + ')'
        jupas_df['School Code'] = str(20212)
        jupas_df['Last Name'] = jupas_df['enname'].str.split(' ').str[0]
        jupas_df['First Name'] = jupas_df['enname'].str.split(' ').str[1:]
        jupas_df['First Name'] = jupas_df['First Name'].apply(lambda l: ' '.join(l))
        jupas_df['Passport No. (Optional)'] = None
        jupas_df['Passport Issuing Country (Optional)'] = None

        jupas_df.rename(columns={'key': 'Group (Optional)',
                                 'classcode': 'Class',
                                 'regno': 'Student No. (Optional)'},
                        inplace=True)

        eapp_df['School Code'] = 190560
        eapp_df['Last Name'] = eapp_df['enname'].str.split(' ').str[0]
        eapp_df['First Name'] = eapp_df['enname'].str.split(' ').str[1:]
        eapp_df['First Name'] = eapp_df['First Name'].apply(lambda l: ' '.join(l))
        eapp_df['Group'] = ''
        eapp_df['Passport No.'] = None
        eapp_df['Passport Issuing Country'] = None

        eapp_column = ['hkid', 'School Code', 'Last Name', 'First Name', 'classcode', 'classno',
                       'Group',	'Passport No.', 'Passport Issuing Country']

        # print(jupas_df.head(5))
        jupas_file_save = os.path.join(self.classlist_home, 'export', 'jupas', self.full_name + '-jupas-import.xlsx')
        jupas_df[jupas_column].to_excel(jupas_file_save, sheet_name='Record', index=False)
        print('\tjupas file: {}'.format(jupas_file_save))

        eapp_file_save = os.path.join(self.classlist_home, 'export', 'jupas', self.full_name + '-eapp-import.xls')
        eapp_df[eapp_column].to_excel(eapp_file_save, sheet_name='Record', index=False, header=False)
        print('\teapp file: {}'.format(eapp_file_save))

        return 0

    def export_jupas_srr(self):

        print('\ngenerate S6 jupas list')
        filter_column = ['key','regno', 'enname', 'chname', 'sex', 'classlevel', 'classcode', 'classno', 'jupas']
        jupas_srr_column = ['jupas', 'regno', 'classlevel', 'classcode', 'classno', 'enname', 'chname', 'sex']

        jupas_srr_df = (self.stuinfo_df[self.stuinfo_df['classlevel'] == 'S6'][filter_column]
                    .applymap(lambda x: re.sub(ILLEGAL_CHARACTERS_RE, '', x) if isinstance(x, str) else x)
                    .sort_values(by=['classlevel', 'classcode', 'classno'], ascending=[True, True, True])
                    )

        # print(jupas_srr_df.head(5))
        jupas_file_save = os.path.join(self.classlist_home, 'export', 'jupas', self.full_name + '-jupas-import.xlsx')
        jupas_srr_df[jupas_srr_column].to_excel(jupas_file_save, sheet_name='Record', index=False)
        print('\tjupas file: {}'.format(jupas_file_save))

        return 0

    def filter_analysis_eca(self):
        eac_header_s123 = ['key', 'regno', 'enname', 'chname', 'sex', 'classlvl', 'classcode', 'classno',
                           'chi', 'eng', 'mth', 'chs', 'hst', 'geo', 'eco', 'isc', 'phy', 'chm', 'bio',
                           'bik', 'lst', 'pth', 'cps', 'via', 'mus', 'ped', 'dte', 'hec',
                           'wm', 'Class Rank', 'Form Rank']

        eac_header_s456 = ['key', 'regno', 'enname', 'chname', 'sex', 'classlvl', 'classcode', 'classno',
                           'chi', 'eng', 'mth', 'm1', 'm2', 'lst',
                           'x1_', 'x1_m', 'x1_k', 'x2_', 'x2_m', 'x2_k', 'x3_', 'x3_m', 'x3_k', 'm_m', 'm_k',
                           'wm', 'DSE   WM', 'X-subj Mean']

        print('\ngenerate student filter list for eca')
        analysis_folder_previous_year = os.path.join(self.home_folder, self.previous_sch_year + 'exam', 'analysis')
        analysis_file_dict = {'s123': os.path.join(analysis_folder_previous_year,
                                                   self.previous_sch_year + '-final-s123-analysis_Final.xlsm'),
                              's456': os.path.join(analysis_folder_previous_year,
                                                   self.previous_sch_year + '-final-s456-analysis_Final.xlsm'),
                              }

        if pathlib.Path(analysis_file_dict['s123']).exists() and pathlib.Path(analysis_file_dict['s456']).exists:
            # 1.    S.1 to S.3 Students who have the rank from 1 to 30 in the overall results of the last academic year.
            # 2.    S.4 Students who got the results with 60% or above in chi eng lst mth
            #       and dse electives subject respectively.
            # 3.    S.5 Students who got the average score of 50% or above in chi eng lst mth
            #       and all dse electives subject.
            s123_rank = 30
            s4_pass_mark = 60
            s5_dse_wm = 50

            save_wb = openpyxl.load_workbook(self.template_file)
            save_file = os.path.join(self.classlist_home, 'export', 'eac', self.sch_year + '-eac-student.xlsx')
            wb_writer = pd.ExcelWriter(save_file, engine='openpyxl')
            wb_writer.book = save_wb

            analysis_s123_df = pd.read_excel(analysis_file_dict['s123'], sheet_name='work')[
                ['regno', 'enname', 'chname', 'sex', 'classlvl', 'classcode', 'classno', 'wm', 'Form Rank']]

            (analysis_s123_df[analysis_s123_df['Form Rank'] <= s123_rank]
                .to_excel(wb_writer, 's123', index=False))

            analysis_s456_df = pd.read_excel(analysis_file_dict['s456'], sheet_name='work')[
                ['key', 'regno', 'enname', 'chname', 'sex', 'classlvl', 'classcode', 'classno',
                 'chi', 'eng', 'mth', 'lst',
                 'x1_', 'x1_m', 'x1_k', 'x2_', 'x2_m', 'x2_k', 'x3_', 'x3_m', 'x3_k',
                 'm_m', 'm_k', 'wm', 'DSE   WM', 'X-subj Mean']]

            s4_df = analysis_s456_df[analysis_s456_df['classlvl'] == 'S4']
            # s4_df.to_excel(wb_writer, 's4o', index=False)
            x1_filter = (s4_df['x1_m'] > s4_pass_mark) | (s4_df['x1_m'].isna())
            x2_filter = (s4_df['x2_m'] > s4_pass_mark) | (s4_df['x2_m'].isna())
            x3_filter = (s4_df['x3_m'] > s4_pass_mark) | (s4_df['x3_m'].isna())
            m_filter = (s4_df['m_m'] > s4_pass_mark) | (s4_df['m_m'].isna())
            (s4_df[(
             (s4_df['chi'] > s4_pass_mark) &
             (s4_df['eng'] > s4_pass_mark) &
             (s4_df['mth'] > s4_pass_mark) &
             (s4_df['lst'] > s4_pass_mark) &
             x1_filter & x2_filter & x3_filter & m_filter
             )]
             .to_excel(wb_writer, 's4', index=False))

            (analysis_s456_df[(analysis_s456_df['classlvl'] == 'S5') & (analysis_s456_df['DSE   WM'] >= s5_dse_wm)][
                 ['regno', 'enname', 'chname', 'sex', 'classlvl', 'classcode', 'classno', 'DSE   WM']]
             .to_excel(wb_writer, 's5', index=False)
             )
            print('\tfile saved: {}'.format(save_file))
            wb_writer.save()
            wb_writer.close()

        else:
            print('{} or {} is missing'.format(analysis_file_dict['s123'], analysis_file_dict['s456']))

    def import_websams(self):

        # 'Promotion Status' takes the following numeric values =
        # 0 - Promoted, 1 - Repeated, 2 - Acc. promoted, 3 - Demoted, 4 - Graduated,
        # 5 - Departed, 6 - Complete S3, 7 - Extended, 8 - Transferred,
        # - Not Assigned

        # promotion_statue = {'promoted': 0, 'repeat': 1, 'graduated': 4,
        #                     'departed': 5, 'transferred': 8, 'not assigned': '-'}

        promotion_header = ['*School ID', '*From School Year', '*From School Level', '*From School Session',
                            '*From Class Level', '*From Class', '*Class Number', '*To School Year', 'To School Level',
                            'To School Session', 'To Class Level', 'To Class', '*Student name', '*Registration No',
                            'Promotion Status']

        promotion_header_dict = dict(zip(promotion_header, range(0, len(promotion_header))))

        student_dict = {'classlevel': 'To Class Level',
                        'classcode': 'To Class',
                        'statue': 'Promotion Status'}

        current_directory = pathlib.Path(self.export_src_folder)
        print('start produce promotion files from', self.export_src_folder)
        # read xls using xlrd
        # loop all files in websams import folder
        self.stuinfo_df['regno2'] = '#' + self.stuinfo_df['regno']
        self.stuinfo_df['statue'].fillna('0', inplace=True)
        stuinfo_dict = self.stuinfo_df.set_index('regno2').to_dict(orient='index')

        for current_file in current_directory.iterdir():
            # PM_1905603320200820_000.xls
            if current_file.name[0:2] == 'PM':

                # current_file_item = {'prefix': 'PM_',
                #                     'school_code': current_file.name[3:9],
                #                     'school_level': current_file.name[9:10],
                #                     'school_session': current_file.name[10:11],
                #                     'dateVal': current_file.name[11:]}

                promotion_file_src = xlrd.open_workbook(current_file)
                copy_sheet = promotion_file_src.sheet_by_index(0)
                temp_wb = xlutils.copy.copy(promotion_file_src)

                # loop through all student in xls import file
                for row in range(1, len(copy_sheet.col_values(0))):

                    cur_regno = copy_sheet.cell(row, promotion_header_dict['*Registration No']).value

                    # check student in all_stu_dict
                    if cur_regno in stuinfo_dict.keys():

                        cur_student = stuinfo_dict[cur_regno]
                        temp_wb.get_sheet(0).write(row, promotion_header_dict['To School Level'], '3')
                        temp_wb.get_sheet(0).write(row, promotion_header_dict['To School Session'], '3')

                        # print(cur_student['classlevel'], cur_student['classcode'], cur_student['statue'])
                        for key, item in student_dict.items():

                            col = promotion_header_dict[item]
                            temp_wb.get_sheet(0).write(row, col, cur_student[key])

                            if key == 'statue':

                                if copy_sheet.cell(row, promotion_header_dict['Promotion Status']).value == '5':
                                    pass

                                elif cur_student[key] == 'repeat':
                                    temp_wb.get_sheet(0).write(row,
                                                               promotion_header_dict['Promotion Status'],
                                                               '1')
                                else:
                                    temp_wb.get_sheet(0).write(row,
                                                               promotion_header_dict['Promotion Status'],
                                                               cur_student[key])

                    else:
                        print('{} ({}) does not exist in the current classlist.'.format(
                            cur_regno,
                            copy_sheet.cell(row, promotion_header_dict['*Student name']).value)
                            )

                save_file = os.path.join(self.export_folder, current_file.name)
                temp_wb.save(save_file)
                # print('\t' + current_file_item['classcode'] + ': ' + save_file)

        print('All import files: {}'.format(self.export_folder))
