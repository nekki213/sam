import os
import pandas as pd
import pathlib
import openpyxl
from openpyxl.styles import Protection

from model.examclass import ExamClass


class ExamRun:

    def __init__(self, exam_year):
        self.location = 'model/'
        self.exam_year = exam_year
        # self.exam_year = '1920'
        self.full_exam = self.exam_year + 'exam'

        # /websams
        self.websams_root_folder = os.path.abspath(os.getcwd())
        # /websams/2021
        self.schyear_home_folder = os.path.join(self.websams_root_folder, self.exam_year)
        # /websams/2021/2021-exam-run.xlsm
        self.exam_run_file = os.path.join(self.schyear_home_folder, self.exam_year + '-exam-run.xlsm')
        # /websams/2021/2021exam   (created by Assessment)
        self.assessment_root_folder = os.path.join(self.schyear_home_folder, self.exam_year + 'exam')
        # /websams/2021/_template
        self.exam_template_folder = os.path.join(self.schyear_home_folder, '_template', 'exam')
        # should copy template to exam_home *first* time

        self.exam_run_df = None
        self.exam_class_list = {}
        self.exam_subject = None
        self.no_of_files = None
        self.exam_run_header = ['ind', 'select', 'key', 'classlevel', 'classcode', 'classtype',
                                'groupcode', 'subject', 'subjcode', 'subjkey', 'path', 'tch',
                                'tch_src', 'room', 'ut1', 'exam1', 'ut2', 'exam2', 'create',
                                'rename', 'basename', 'create_folder', 'rename_folder']
        self.student_sheet_header = ['key', 'regno', 'enname', 'chname', 'sex',
                                     'house', 'classlevel', 'classcode', 'classno', 'old classlevel',
                                     'old classcode', 'old classno', 'statue',
                                     'x1', 'x2', 'x3', 'm', 'mth', 'mth_tch', 'DH']
        # need a drop sheet header
        self.drop_sheet_header = []
        self.stuinfo_df = None
        self.drop_df = None
        self.promotion = None
        self.assessment_dict = {}
        self.ct_dict = {}
        # 145	O	s3bchi	s3	s3b
        # class	s3b	chi	080	s3bchi
        # tch	CS	CS	rmC
        # O	O	O	O
        # ut1	H	s3bchi.xlsx
        # ut1\chi\1920ut1s3bchi.xlsx	exam1\CS

    def load_stuinfo(self):
        print('please update stuinfo sheet in exam-run before running.')
        print('load stuinfo sheet from exam-run.')
        self.stuinfo_df = pd.read_excel(self.exam_run_file, sheet_name='stuinfo')
        return len(self.stuinfo_df)

    def load_ct_to_dict(self):
        # load class teachers from exam-run.xlsm
        if self.ct_dict:
            pass
        else:
            ct_df = pd.read_excel(self.exam_run_file, sheet_name='ct')
            ct_df['classcode2'] = ct_df['classcode']
            ct_df.set_index('classcode2', inplace=True)
            self.ct_dict = ct_df[['ct1', 'ct2']].to_dict('index')
        return self.ct_dict

    def markfile_name(self, class_type, exam_type, class_level, group_code, subject, teacher):
        if class_type == 'group':
            temp = '_'.join([self.exam_year, exam_type, class_level, group_code, subject, teacher.lower()]) + '.xlsx'
            temp = self.exam_year + exam_type + class_level + group_code + subject + '.xlsx'

        else:
            temp = '_'.join([self.exam_year, exam_type, group_code, subject, teacher.lower()]) + '.xlsx'
            temp = self.exam_year + exam_type + group_code + subject + '.xlsx'

        return temp

    def load_drop(self):
        print('load stuinfo sheet.')
        self.drop_df = pd.read_excel(self.exam_run_file, sheet_name='drop')
        return len(self.drop_df)

    def create_exam_template(self):
        import openpyxl
        # should run when exam files are created.
        print('exam template will be create.')
        print('stuinfo will be copied to: s123-analysis.xlsm, s456-analysis.xlsm, view-all.xlsm')
        s123_template = os.path.join(self.exam_template_folder, 's123-analysis.xlsm')
        s456_template = os.path.join(self.exam_template_folder, 's456-analysis.xlsm')
        view_template = os.path.join(self.exam_template_folder, 'view-all.xlsm')

        exam_run_wb = openpyxl.load_workbook(filename=self.exam_run_file)
        stuinfo_ws = exam_run_wb['stuinfo']
        run_ws = exam_run_wb['run']

        s123_wb = openpyxl.load_workbook(filename=s123_template, keep_vba=True)
        s123_stuinfo_ws = s123_wb['stuinfo']

        s456_wb = openpyxl.load_workbook(filename=s123_template, keep_vba=True)
        s456_stuinfo_ws = s456_wb['stuinfo']

        view_wb = openpyxl.load_workbook(filename=view_template, keep_vba=True)
        view_stuinfo_ws = view_wb['stuinfo']
        view_index_ws = view_wb['index']

        s123_wb_save = os.path.join(self.exam_template_folder, self.exam_year + '-s123-analysis.xlsm')
        s456_wb_save = os.path.join(self.exam_template_folder, self.exam_year + '-s456-analysis.xlsm')
        view_wb_save = os.path.join(self.exam_template_folder, self.exam_year + '-view-all.xlsm')

        s123_wb.save(filename=s123_wb_save)
        s123_wb.close()

        s456_wb.save(filename=s456_wb_save)
        s456_wb.close()

        view_wb.save(filename=view_wb_save)
        view_wb.close()

    def load_run(self):
        self.exam_class_list = {}
        print('load run sheet.')
        self.exam_run_df = pd.read_excel(self.exam_run_file, sheet_name='run')
        # self.exam_run_header = self.exam_run_df.columns.values
        temp_dict = self.exam_run_df.to_dict(orient='records')

        for item in temp_dict:
            # print(item['Key'])
            self.exam_class_list[item['key']] = ExamClass(examyear=self.exam_year,
                                                          ind=item['ind'],
                                                          select=item['select'],
                                                          key=item['key'],
                                                          classlevel=item['classlevel'],
                                                          classcode=item['classcode'],
                                                          classtype=item['classtype'],
                                                          groupcode=item['groupcode'],
                                                          subject=item['subject'],
                                                          subj_code=item['subjcode'],
                                                          subj_key=item['subjkey'],
                                                          path=item['path'],
                                                          tch=item['tch'],
                                                          teacher=item['tch_src'],
                                                          room=item['room'],
                                                          ut1=item['ut1'],
                                                          exam1=item['exam1'],
                                                          ut2=item['ut2'],
                                                          exam2=item['exam2'],
                                                          create=item['create'],
                                                          rename=item['rename'],
                                                          basename=item['basename'],
                                                          create_folder=item['create_folder'],
                                                          rename_folder=item['rename_folder'],
                                                          )

    def load_assessment(self, assessment):
        self.assessment_dict = {}
        print('load run sheet for {}.'.format(assessment))
        assessment_df = pd.read_excel(self.exam_run_file, sheet_name='run')
        print(len(assessment_df))
        assessment_filter = assessment_df[assessment] == 'O'
        assessment_df = assessment_df[assessment_filter]
        temp_dict = assessment_df.to_dict('records')

        for item in temp_dict:
            # print(item['Key'])
            self.assessment_dict[item['key']] = ExamClass(examyear=self.exam_year,
                                                          ind=item['ind'],
                                                          select=item['select'],
                                                          key=item['key'],
                                                          classlevel=item['classlevel'],
                                                          classcode=item['classcode'],
                                                          classtype=item['classtype'],
                                                          groupcode=item['groupcode'],
                                                          subject=item['subject'],
                                                          subj_code=item['subjcode'],
                                                          subj_key=item['subjkey'],
                                                          path=item['path'],
                                                          tch=item['tch'],
                                                          teacher=item['tch_src'],
                                                          room=item['room'],
                                                          ut1=item['ut1'],
                                                          exam1=item['exam1'],
                                                          ut2=item['ut2'],
                                                          exam2=item['exam2'],
                                                          create=item['create'],
                                                          rename=item['rename'],
                                                          basename=item['basename'],
                                                          create_folder=item['create_folder'],
                                                          rename_folder=item['rename_folder'],
                                                          )

    def create(self):
        # design a weight checking later
        markfile_col_list = ['regno', 'enname', 'chname', 'sex', 'classlevel', 'classcode', 'classno']
        markfile_col = {key: i + 1 for i, key in enumerate(markfile_col_list)}
        lock_range = {'chi': {'ut1': [15, 1], 'daily1': [15, 9], 'exam1': [16, 4],
                              'ut2': [15, 1], 'daily2': [15, 9], 'exam2': [16, 4], },
                      'eng': {'ut1': [8, 6], 'daily1': [0, 0], 'exam1': [9, 12],
                              'ut2': [8, 6], 'daily2': [0, 0], 'exam2': [9, 12], },
                      'oth': {'ut1': [11, 1], 'daily1': [0, 0], 'exam1': [10, 2],
                              'ut2': [11, 1], 'daily2': [0, 0], 'exam2': [10, 2], },
                      }

        print('create files for', self.exam_year)
        print('load student info and exam master file')
        # load student subjects as dataframe
        self.load_stuinfo()
        # print(self.stuinfo_df)
        self.load_run()
        # print(len(self.exam_class_list))

        # build folders first
        markfile_src_root_folder = os.path.join(self.assessment_root_folder, 'markfile_src')
        os.makedirs(markfile_src_root_folder, exist_ok=True)

        for folder in ['ut1', 'exam1', 'ut2', 'exam2', 'mock']:
            folder_to_create = os.path.join(markfile_src_root_folder, folder)
            os.makedirs(folder_to_create, exist_ok=True)

        num_of_markfiles = 0
        for key, exam_class in self.exam_class_list.items():

            class_type = exam_class.class_type
            class_level = exam_class.classlevel     # s6
            group_code = exam_class.groupcode       # s6x / x1 / x2 / x3
            subject = exam_class.subject
            exam_type = exam_class.create
            classcode = exam_class.classcode.upper()
            teacher = exam_class.tch

            # print(exam_class)
            # print(exam_class.class_type)
            # a classlevel filter for restrictive test run
            if exam_class.select == 'O':

                if exam_class.classlevel.lower() in ['s1', 's2', 's3', 's4', 's5', 's6']:

                    if exam_class.class_type == 'class':
                        # filter dte/hec students for s2 and s3
                        if exam_class.subject.lower() in ['dte', 'hec'] and exam_class.classlevel.lower() in ['s2', 's3']:
                            class_df = self.stuinfo_df[(self.stuinfo_df.classcode.isin([classcode]))
                                & (self.stuinfo_df['DH'] == exam_class.subject)][markfile_col_list]
                            class_dict = class_df.to_dict('records')

                        elif exam_class.subject.lower() in ['pedm', 'pedf']:
                            sex_code = exam_class.subject.lower()[-1]
                            class_df = self.stuinfo_df[(self.stuinfo_df.classcode.isin([classcode]))
                                & (self.stuinfo_df['sex'] == sex_code.upper())][markfile_col_list]
                            class_dict = class_df.to_dict('records')

                        else:
                            class_df = self.stuinfo_df[self.stuinfo_df.classcode.isin([classcode])][markfile_col_list]
                            class_dict = class_df.to_dict('records')

                    elif exam_class.class_type == 'group':
                        # print(exam_class.groupcode, exam_class.subject.lower())
                        if exam_class.groupcode[0] == 'x':
                            class_df = self.stuinfo_df[(self.stuinfo_df['classlevel'] == exam_class.classlevel.upper())
                                & (self.stuinfo_df[exam_class.groupcode] == exam_class.subject.lower())][markfile_col_list]
                            class_dict = class_df.to_dict('records')
                            # print(class_dict)

                        elif exam_class.groupcode[0] == 'g':
                            # did not decide for the old s6 mth/m1/m2 split classes
                            # good for new m1 as x1 elective
                            # print(subject)
                            # distinguish m1, m2, mth when 'g'
                            # so that ok for non-x-subject setting of m1/m2
                            if exam_class.subject == 'mth':
                                class_df = self.stuinfo_df[(self.stuinfo_df['classlevel'] == exam_class.classlevel.upper())
                                    & (self.stuinfo_df[exam_class.subject.lower()] == exam_class.groupcode)][markfile_col_list]
                                class_dict = class_df.to_dict('records')
                            elif exam_class.subject in ['m1', 'm2']:
                                class_df = self.stuinfo_df[(self.stuinfo_df['classlevel'] == exam_class.classlevel.upper())
                                    & (self.stuinfo_df['m'] == exam_class.subject.lower())][markfile_col_list]
                                class_dict = class_df.to_dict('records')
                            else:
                                class_df = []

                    else:
                        print('Error: not class nor group!')

                    # print(class_dict)
                    # load mark file template
                    markfile_template_folder = os.path.join(self.websams_root_folder,
                                                            'setup',
                                                            str(self.exam_year) + 'files')

                    # print(markfile_template_folder, exam_class.subject, exam_class.classlevel.lower(), exam_type)
                    markfile_template = os.path.join(markfile_template_folder,
                                                     exam_class.subject,
                                                     exam_class.classlevel.lower() + 'x.xlsx')
                    # websams\1920\1920exam\markfile_src\{ exam }
                    # print(self.home_folder, self.full_exam, 'markfile_src', exam_type, markfile_template)
                    markfile_src_folder = os.path.join(markfile_src_root_folder, exam_type)

                    # print(markfile_src_folder)
                    markfile_wb = openpyxl.load_workbook(filename=markfile_template)
                    websams_ws = markfile_wb['websams']
                    setup_ws = markfile_wb['setup']

                    # set the file key on setup sheet
                    setup_ws.cell(column=1, row=2).value = key
                    setup_ws.cell(column=4, row=1).value = exam_class.subject.lower()
                    if exam_class.groupcode in ['x1', 'x2', 'x3', 'g1', 'g2']:
                        setup_ws.cell(column=3, row=1).value = exam_class.groupcode
                    # write student data on websams sheet by loop through class list dict
                    # ** need to ensure no strange white spaces in names
                    currow = 5
                    col_taken = 45
                    for stu in class_dict:
                        # print(stu)
                        for item in markfile_col:
                            # print(item)
                            websams_ws.cell(column=markfile_col[item], row=currow).value = stu[item]
                        websams_ws.cell(column=col_taken, row=currow).value = 'O'
                        currow += 1

                    # loop all sheets in workbook
                    # to lock cells
                    for sheet in markfile_wb.sheetnames:
                        if sheet in ['ut1', 'daily1', 'exam1', 'ut2', 'daily2', 'exam2']:
                            # set unlocked cells first
                            if exam_class.subject.lower() in ['chi', 'eng']:
                                col_start = lock_range[exam_class.subject.lower()][sheet][0]
                                repeat = lock_range[exam_class.subject.lower()][sheet][1]
                            else:
                                col_start = lock_range['oth'][sheet][0]
                                repeat = lock_range['oth'][sheet][1]
                            # loop the lock_range range
                            # print(sheet, colStart, repeat)
                            for r2 in range(5, 49):
                                for c2 in range(col_start, col_start + repeat):
                                    markfile_wb[sheet].cell(column=c2, row=r2).protection = Protection(locked=False,
                                                                                                       hidden=False)
                        markfile_wb[sheet].protection.password = 'css'
                        markfile_wb[sheet].protection.sheet = True
                        markfile_wb[sheet].protection.enable()

                    # build teacher and subject folder first
                    # for check and create if necessary
                    teacher_folder = os.path.join(markfile_src_folder, 'teacher')
                    subject_folder = os.path.join(markfile_src_folder, 'subject')
                    # os.makedirs(teacher_folder, exist_ok=True)
                    os.makedirs(subject_folder, exist_ok=True)

                    # check folder -> teacher or subject
                    # set the file save folder and filename
                    # override exam_type for s6 mock
                    if exam_class.classlevel.lower() == 's6' and exam_type == 'mock':
                        exam_type = 'mock'

                    mark_filename = self.markfile_name(class_type, exam_type, class_level, group_code, subject, teacher)
                    file_save_folder = os.path.join(subject_folder, exam_class.subject)
                    os.makedirs(file_save_folder, exist_ok=True)
                    file_save = os.path.join(file_save_folder, mark_filename)

                    # from https://groups.google.com/forum/#!topic/openpyxl-users/Y9_iSeTi3bM
                    #
                    #    FWIW this is what I used to work around something similar in a file with
                    #    chart sheets and invisible worksheets with data:
                    #
                    #    wb.views[0].firstSheet = 1
                    #
                    #    I'd love to be able to explain this but I can't as the specification says
                    #    nothing about grouping worksheets. :-/ If anyone can come up with
                    #    something suitable I'd love to add it to the documentation!
                    #
                    # it just does not work.
                    # k = markfile_wb.views[0].activeTab
                    # markfile_wb.views[0].firstSheet = k
                    # print(markfile_wb.views[0].firstSheet, markfile_wb.views[0].activeTab)
                    num_of_markfiles += 1
                    print('\t#{} created: {}'.format(str(num_of_markfiles), file_save))
                    markfile_wb.save(filename=file_save)

        print('total no of markfiles: {}'.format(num_of_markfiles))

    def unlock_all(self):
        # from exam_type (= exam_folder)
        # merge_folder = self.
        print('current exam:', self.merge_folder)
        print('never stop when writing excel files.')
        print('otherwise files will be corrupted.')
        # loop current merge folder
        # unlock and save

        weight = {'s1': {'mth': [0.7, 0.3, 1, 0, 0],
                         'lst': [0.7, 0.3, 1, 0, 0],
                         'isc': [0.5, 0.5, 0.25, 0, 0.75],  ##
                         'chs': [0.7, 0.3, 1, 0, 0],
                         'geo': [0.7, 0.3, 1, 0, 0],
                         'bik': [0.7, 0.3, 1, 0, 0],
                         },
                  's2': {'mth': [0.7, 0.3, 1, 0, 0],
                         'lst': [0.7, 0.3, 1, 0, 0],
                         'isc': [0.7, 0.3, 1, 0, 0],
                         'chs': [0.5, 0.5, 0.25, 0, 0.75],
                         'hst': [0.7, 0.3, 1, 0, 0],
                         'geo': [0.7, 0.3, 1, 0, 0],
                         'bik': [0.7, 0.3, 1, 0, 0],
                         },
                   's3': {'mth': [0.5, 0.5, 0.25, 0, 0.75],
                          'lst': [0.5, 0.5, 1, 0, 0],
                          'phy': [0.5, 0.5, 0.25, 0, 0.75],
                          'chm': [0.5, 0.5, 0.25, 0, 0.75],
                          'bio': [0.5, 0.5, 0.25, 0, 0.75],
                          'chs': [0.5, 0.5, 0.25, 0, 0.75],
                          'hst': [0.5, 0.5, 0.25, 0, 0.75],
                          'geo': [0.5, 0.5, 0.25, 0, 0.75],
                          'eco': [0.5, 0.5, 0.25, 0, 0.75],
                          'bik': [0.5, 0.5, 0.25, 0, 0.75],
                          },
                   's4': {'mth': [0.7, 0.3, 1, 0, 0],
                          'm1': [0.7, 0.3, 1, 0, 0],
                          'm2': [0.7, 0.3, 1, 0, 0],
                          'lst': [0.5, 0.5, 0, 0.25, 0.75],
                          'phy': [0.7, 0.3, 1, 0, 0],
                          'chm': [0.5, 0.5, 0.25, 0, 0.75],
                          'bio': [0.5, 0.5, 0.25, 0, 0.75],
                          'chs': [0.5, 0.5, 0.25, 0, 0.75],
                          'hst': [0.7, 0.3, 1, 0, 0],
                          'geo': [0.7, 0.3, 1, 0, 0],
                          'eco': [0.7, 0.3, 1, 0, 0],
                          'ict': [0.5, 0.5, 0.25, 0, 0.75],
                          'baf': [0.5, 0.5, 0.25, 0, 0.75],
                          },
                   's5': {'mth': [0.7, 0.3, 1, 0, 0],
                          'm1': [0.5, 0.5, 0.25, 0, 0.75],
                          'm2': [0.5, 0.5, 0.25, 0, 0.75],
                          'lst': [0.7, 0.3, 1, 0, 0],
                          'phy': [0.7, 0.3, 1, 0, 0],
                          'chm': [0.5, 0.5, 0.25, 0, 0.75],
                          'bio': [0.5, 0.5, 0.25, 0, 0.75],
                          'chs': [0.7, 0.3, 1, 0, 0],
                          'hst': [0.5, 0.5, 0.25, 0, 0.75],
                          'geo': [0.5, 0.5, 0.25, 0, 0.75],
                          'eco': [0.7, 0.3, 1, 0, 0],
                          'ict': [0.7, 0.3, 1, 0, 0],
                          'baf': [0.7, 0.3, 1, 0, 0],
                          },
                   }

        current_directory = pathlib.Path(self.merge_folder)
        for subject_dir in current_directory.iterdir():
            # only loop into sub-folder of merge
            if subject_dir.is_dir():
                # produce a list of files to loop using os.scandir
                print(subject_dir)
                with os.scandir(str(subject_dir)) as listOfEntries:
                    # loop all files in each subject folder
                    for entry in listOfEntries:
                        # check if files vs directory
                        if entry.is_file():
                            # only open xlsx file
                            if entry.name[-4:] == 'xlsx':
                                n = len(self.exam_year + self.exam_filetype)
                                exam_type = entry.name[0:n]
                                classlevel = entry.name[n:(n+2)]
                                k = len(entry.name[n:])
                                base_markfile = entry.name[n:]
                                if base_markfile[2] in ['x', 'g']:
                                    subject = base_markfile[4:-5]
                                else:
                                    subject = base_markfile[3:-5]

                                # subject = (base_markfile[4:-5] if base_markfile[2] == 'x' else base_markfile[3:-5])
                                # print(k, base_markfile[4:k-5], base_markfile[3:k-5])
                                # need subject here
                                if subject.lower() in ['lst']:
                                    markfile_wb = openpyxl.load_workbook(filename=entry.path)
                                    print(base_markfile)

                                    for sheet in markfile_wb.sheetnames:
                                        markfile_wb[sheet].protection.password = 'css'
                                        markfile_wb[sheet].protection.sheet = False
                                        markfile_wb[sheet].protection.disable()

                                        term1_col = 2
                                        term2_col = 3
                                        daily2_col = 10
                                        ut2_col = 11
                                        exam2_col = 12

                                        # if sheet == 'setup' and subject not in ['chi', 'eng']:
                                        #    markfile_wb['websams'].cell(row=2, column=21).value = 0
                                        #    print(classlevel, subject, weight[classlevel][subject])
                                        #    markfile_wb['setup'].cell(row=4, column=term1_col).value = weight[classlevel][subject][0]
                                        #    markfile_wb['setup'].cell(row=4, column=term2_col).value = weight[classlevel][subject][1]
                                        #    markfile_wb['setup'].cell(row=4, column=daily2_col).value = weight[classlevel][subject][2]
                                        #    markfile_wb['setup'].cell(row=4, column=ut2_col).value = weight[classlevel][subject][3]
                                        #    markfile_wb['setup'].cell(row=4, column=exam2_col).value = weight[classlevel][subject][4]

                                    markfile_wb.save(entry.path)

    # rename from cur_exam_type
    def rename(self, past_exam, next_exam, lock_state=True):

        lock_range = {'chi': {'ut1': [15, 1], 'daily1': [15, 12], 'exam1': [16, 12],
                              'ut2': [15, 1], 'daily2': [15, 12], 'exam2': [16, 12]},
                      'eng': {'ut1': [8, 6], 'daily1': [0, 0], 'exam1': [9, 14],
                              'ut2': [8, 6], 'daily2': [0, 0], 'exam2': [9, 14]},
                      'oth': {'ut1': [11, 1], 'daily1': [0, 0], 'exam1': [10, 2],
                              'ut2': [11, 1], 'daily2': [0, 0], 'exam2': [10, 2]},
                      }

        markfile_src_folder = os.path.join(self.assessment_root_folder, 'markfile_src', next_exam)

        # from exam_type (= exam_folder)
        # also load drop from master run
        # unlock websams and update subject for S5

        # set path 2021exam\ut1\(subject)
        past_exam_folder = os.path.join(self.assessment_root_folder, past_exam, 'merge')
        print('past exam:', past_exam_folder)

        # load master exam file
        exam_run_df = pd.read_excel(self.exam_run_file, sheet_name='run')
        exam_run_df = exam_run_df[(exam_run_df[past_exam] == 'O') & (exam_run_df[next_exam] == 'O')]
        file_rename_list = exam_run_df['basename'].to_list()
        print(file_rename_list)
        # excel_file_tch_dict = dict(zip(exam_run_df.file, exam_run_df.TchV))
        # form drop student dict
        # need more work on dict
        # drop_header = ['term', 'regno', 'chname', 'classlevel', 'classcode', 'classno', 'subject', 'basefile']
        # drop_df = pd.read_excel(self.exam_master_file, sheet_name='drop')
        # drop_df = drop_df[drop_header]
        # drop_file_dict = dict(zip(drop_df.basefile, drop_df.to_dict(orient='records')))
        # print(drop_file_dict)
        # loop current merge folder
        # and save to next exam
        current_directory = pathlib.Path(past_exam_folder)
        for subject_dir in current_directory.iterdir():
            # only loop into sub-folder of merge
            if subject_dir.is_dir():
                # produce a list of files to loop using os.scandir
                with os.scandir(str(subject_dir)) as listOfEntries:
                    # loop all files in each subject folder
                    for entry in listOfEntries:
                        # check if files vs directory
                        if entry.is_file():
                            # only open xlsx file
                            if entry.name[-4:] == 'xlsx':
                                n = len(self.exam_year + past_exam)
                                exam_type = entry.name[0:n]
                                classlevel = entry.name[n:(n+2)]
                                base_markfile = entry.name[n:]
                                base_markfile2 = entry.name[n:-5]
                                print(base_markfile, base_markfile2)
                                # subject = (base_markfile[4:-5] if base_markfile[2] == 'x' else base_markfile[3:-5])
                                if base_markfile[2] in ['x', 'g']:
                                    subject = base_markfile[4:-5]
                                else:
                                    subject = base_markfile[3:-5]

                                if base_markfile in file_rename_list:
                                    print(entry.name, subject, classlevel, past_exam)
                                    # tch = excel_file_tch_dict[base_markfile]
                                    # markfile_save_folder = self.exam_folder + next_exam + sep \
                                    #    + 'markfile_src' + sep + 'rename' + sep + tch + sep
                                    markfile_save_folder = os.path.join(self.exam_home,
                                                                        'markfile_src', next_exam,
                                                                        'rename', subject)
                                    print(markfile_save_folder)
                                    if not os.path.isdir(markfile_save_folder):
                                        os.makedirs(markfile_save_folder)

                                    markfile_save = os.path.join(markfile_save_folder,
                                                                 self.exam_year + next_exam + base_markfile)
                                    print(exam_type, base_markfile, next_exam, markfile_save)
                                    markfile_wb = openpyxl.load_workbook(filename=entry.path)

                                    # check drop students for nss elective
                                    # if base_markfile2 in drop_file_dict.keys():
                                    #    for row in range(5, 49):
                                    #        regno = markfile_wb['websams'].cell(row=row, column=1).value
                                    #        if regno == drop_file_dict[base_markfile2]['regno']:
                                    #            print(drop_file_dict[base_markfile2]['chname'])
                                    #            markfile_wb['websams'].cell(row=row, column=45).value = ''

                                    # filter out all previous exams
                                    # rewrite lock state for the cells in those ws
                                    # do not lock future exam

                                    locked_sheet_state = {'ut1': True, 'daily1': True, 'exam1': True,
                                                          'ut2': True, 'daily2': False, 'exam2': False,
                                                          }
                                    for ws in [x for x in locked_sheet_state.keys() if locked_sheet_state[x]]:
                                        # set lock_range cells first
                                        if subject.lower() in ['chi', 'eng']:
                                            col_start = lock_range[subject.lower()][ws][0]
                                            repeat = lock_range[subject.lower()][ws][1]
                                        else:
                                            col_start = lock_range['oth'][ws][0]
                                            repeat = lock_range['oth'][ws][1]
                                        # loop the lock_range range
                                        # print(sheet, colStart, repeat)
                                        # sheet_state = locked_sheet_state[ws]

                                        for r2 in range(5, 49):
                                            for c2 in range(col_start, col_start + repeat):
                                                markfile_wb[ws].cell(column=c2, row=r2).protection = Protection(
                                                    locked=True,
                                                    hidden=False)

                                    # loop all ws to update lock state of each ws with password
                                    for ws in markfile_wb.sheetnames:
                                        markfile_wb[ws].protection.password = 'css'
                                        markfile_wb[ws].protection.sheet = False
                                        if lock_state:
                                            markfile_wb[ws].protection.enable()
                                        else:
                                            markfile_wb[ws].protection.disable()

                                    markfile_wb.save(markfile_save)
                                    markfile_wb.close()

    def lock_markfile_2(self, next_exam):

        lock_range = {'chi': {'ut1': [15, 1], 'daily1': [15, 9], 'exam1': [16, 4],
                              'ut2': [15, 1], 'daily2': [15, 9], 'exam2': [16, 4], },
                      'eng': {'ut1': [8, 6], 'daily1': [0, 0], 'exam1': [9, 12],
                              'ut2': [8, 6], 'daily2': [0, 0], 'exam2': [9, 12], },
                      'oth': {'ut1': [11, 1], 'daily1': [0, 0], 'exam1': [10, 2],
                              'ut2': [11, 1], 'daily2': [0, 0], 'exam2': [10, 2], },
                      }

        markfile_save_folder = os.path.join(self.assessment_root_folder, 'markfile_src', next_exam, 'rename')

        # from exam_type (= exam_folder)
        # also load drop from master run
        # unlock websams and update subject for S5
        # set path 2021exam\ut1\(subject)

        # load master exam file
        current_directory = pathlib.Path(markfile_save_folder)
        for subject_dir in current_directory.iterdir():
            # only loop into sub-folder of merge
            if subject_dir.is_dir():
                # produce a list of files to loop using os.scandir
                with os.scandir(str(subject_dir)) as listOfEntries:
                    # loop all files in each subject folder
                    for entry in listOfEntries:
                        # check if files vs directory
                        if entry.is_file():
                            # only open xlsx file
                            if entry.name[-4:] == 'xlsx':
                                n = len(self.exam_year + next_exam)
                                exam_type = entry.name[0:n]
                                classlevel = entry.name[n:(n+2)]
                                base_markfile = entry.name[n:]
                                base_markfile2 = entry.name[n:-5]
                                print(base_markfile, base_markfile2)
                                # subject = (base_markfile[4:-5] if base_markfile[2] == 'x' else base_markfile[3:-5])
                                if base_markfile[2] in ['x', 'g']:
                                    subject = base_markfile[4:-5]
                                else:
                                    subject = base_markfile[3:-5]

                                print(markfile_save_folder)
                                markfile_wb = openpyxl.load_workbook(filename=entry.path)

                                # check drop students for nss elective
                                # if base_markfile2 in drop_file_dict.keys():
                                #    for row in range(5, 49):
                                #        regno = markfile_wb['websams'].cell(row=row, column=1).value
                                #        if regno == drop_file_dict[base_markfile2]['regno']:
                                #            print(drop_file_dict[base_markfile2]['chname'])
                                #            markfile_wb['websams'].cell(row=row, column=45).value = ''

                                # need to unlock and set active ws
                                # need to lock past exam ws
                                # loop all sheets in workbook
                                # to lock cells
                                # unlock files for current exam

                                locked_sheet_state = {'ut1': True, 'daily1': False, 'exam1': False,
                                                      'ut2': False, 'daily2': False, 'exam2': False,
                                                      }

                                for ws in markfile_wb.sheetnames:
                                    if ws in locked_sheet_state.keys():
                                        # set lock_range cells first
                                        if subject.lower() in ['chi', 'eng']:
                                            col_start = lock_range[subject.lower()][ws][0]
                                            repeat = lock_range[subject.lower()][ws][1]
                                        else:
                                            col_start = lock_range['oth'][ws][0]
                                            repeat = lock_range['oth'][ws][1]
                                        # loop the lock_range range
                                        # print(sheet, colStart, repeat)
                                        sheet_state = locked_sheet_state[ws]
                                        for r2 in range(5, 49):
                                            for c2 in range(col_start, col_start + repeat):
                                                markfile_wb[ws].cell(column=c2, row=r2).protection = Protection(
                                                    locked=sheet_state,
                                                    hidden=False)
                                    # else:
                                        # print('undefined sheet!')
                                    markfile_wb[ws].protection.password = 'css'
                                    markfile_wb[ws].protection.sheet = True
                                    markfile_wb[ws].protection.enable()

                                markfile_wb.save(entry.path)
                                markfile_wb.close()

