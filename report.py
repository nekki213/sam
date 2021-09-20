import pandas as pd
import numpy as np
import zipfile
import openpyxl
import os


'''
    1.  unzip reportP files of each form
    2.  load reportP files (student, score, conduct) into pandas and merge
    3.  copy sheets from exam-run and a template
    4.  filter df from (2) and write (using openpyxl) to the template -> sch-rpt excel
    5.  filter df from (2) to generate all ranks
    6.  filter df from (2) for record cards
    
'''


def converter(column: list):
    return {col: str for col in range(len(column))}


class ReportStudent:
    def __init__(self):
        self.name = 'student'
        self.filename = 'STUDENT.xls'
        self.column = [
            'REGNO', 'HKID', 'STRN', 'CHNAME', 'ENNAME', 'SEX', 'DOB',
            'DEPARTURETERM', 'SCHYEAR', 'SCHLEVEL', 'SCHSESSION', 'SCHNAME',
            'CLASSLEVEL', 'CLASS', 'SUBJGRPCODE', 'SUBJGRPDESC', 'STREAM', 'CLASSNO',
            'NEXTYRCLS', 'PROMOTESTATUS', 'ENPROMOTE', 'CHPROMOTE', 'NUMBERSTUDENT',
            'STAFFCODE', 'ENSTAFFNAME', 'CHSTAFFNAME',
            'EXTRA_1', 'EXTRA_2', 'EXTRA_3', 'EXTRA_4', 'EXTRA_5',
            'EXTRA_6', 'EXTRA_7', 'EXTRA_8', 'EXTRA_9', 'EXTRA_10',
            'EXTRA_11', 'EXTRA_12', 'EXTRA_13', 'EXTRA_14', 'EXTRA_15',
            'EXTRA_16', 'EXTRA_17', 'EXTRA_18', 'EXTRA_19', 'EXTRA_20',
            'ECASERVCODE_1', 'ENECASERVDESC_1', 'CHECASERVDESC_1',
            'ECASERVCODE_2', 'ENECASERVDESC_2', 'CHECASERVDESC_2',
            'ECASERVCODE_3', 'ENECASERVDESC_3', 'CHECASERVDESC_3',
            'ECASERVCODE_4', 'ENECASERVDESC_4', 'CHECASERVDESC_4',
            'ECASERVCODE_5', 'ENECASERVDESC_5', 'CHECASERVDESC_5',
            'ECASERVCODE_6', 'ENECASERVDESC_6', 'CHECASERVDESC_6',
            'ECASERVCODE_7', 'ENECASERVDESC_7', 'CHECASERVDESC_7',
            'ECASERVCODE_8', 'ENECASERVDESC_8', 'CHECASERVDESC_8',
            'ECASERVCODE_9', 'ENECASERVDESC_9', 'CHECASERVDESC_9',
            'ECASERVCODE_10', 'ENECASERVDESC_10', 'CHECASERVDESC_10',
            'ENREMARK_A', 'CHREMARK_A', 'ENREMARK_B', 'CHREMARK_B',
            'ENREMARK_C', 'CHREMARK_C', 'ENREMARK_D', 'CHREMARK_D',
            'ENREMARK_E', 'CHREMARK_E', 'ENREMARK_F', 'CHREMARK_F',
            'ENREMARK_G', 'CHREMARK_G', 'ENREMARK_H', 'CHREMARK_H',
            'ENREMARK_I', 'CHREMARK_I', 'ENREMARK_J', 'CHREMARK_J',
            'ENREMARK_K', 'CHREMARK_K',
            'ANPENDESC', 'ANPCHDESC',
            'SCORE_ATA1', 'OM_ATA1', 'ENCOMMENT_ATA1', 'CHCOMMENT_ATA1', 'ANP_ATA1', 'ATT_ATA1',
            'SCORE_ATA2', 'OM_ATA2', 'ENCOMMENT_ATA2', 'CHCOMMENT_ATA2', 'ANP_ATA2', 'ATT_ATA2',
            'SCORE_ATA3', 'OM_ATA3', 'ENCOMMENT_ATA3', 'CHCOMMENT_ATA3', 'ANP_ATA3', 'ATT_ATA3',
            'SCORE_ATA4', 'OM_ATA4', 'ENCOMMENT_ATA4', 'CHCOMMENT_ATA4', 'ANP_ATA4', 'ATT_ATA4',
            'SCORE_ATA5', 'OM_ATA5', 'ENCOMMENT_ATA5', 'CHCOMMENT_ATA5', 'ANP_ATA5', 'ATT_ATA5',
            'SCORE_ATA6', 'OM_ATA6', 'ENCOMMENT_ATA6', 'CHCOMMENT_ATA6', 'ANP_ATA6', 'ATT_ATA6',
            'SCORE_ATA7', 'OM_ATA7', 'ENCOMMENT_ATA7', 'CHCOMMENT_ATA7', 'ANP_ATA7', 'ATT_ATA7',
            'SCORE_ATA8', 'OM_ATA8', 'ENCOMMENT_ATA8', 'CHCOMMENT_ATA8', 'ANP_ATA8', 'ATT_ATA8',
            ]
        self.column_filter = [
            'REGNO', 'HKID', 'CHNAME', 'ENNAME', 'SEX', 'DOB', 'SCHYEAR',
            'CLASSLEVEL', 'CLASS', 'CLASSNO', 'ANPCHDESC',
            'SCORE_ATA1', 'OM_ATA1', 'ANP_ATA1', 'ATT_ATA1',    # ut1
            'SCORE_ATA2', 'OM_ATA2', 'ANP_ATA2', 'ATT_ATA2',    # exam1
            'SCORE_ATA3', 'OM_ATA3', 'ANP_ATA3', 'ATT_ATA3',    # total1
            'SCORE_ATA4', 'OM_ATA4', 'ANP_ATA4', 'ATT_ATA4',    # ut2
            'SCORE_ATA5', 'OM_ATA5', 'ANP_ATA5', 'ATT_ATA5',    # exam2
            'SCORE_ATA6', 'OM_ATA6', 'ENCOMMENT_ATA6', 'CHCOMMENT_ATA6', 'ANP_ATA6', 'ATT_ATA6',    # total2
            'SCORE_ATA7', 'OM_ATA7', 'ANP_ATA7', 'ATT_ATA7',  # final
            # 'SCORE_ATA8', 'OM_ATA8', 'ANP_ATA8', 'ATT_ATA8',    # final
            ]
        self.column_to_view = {
            'CLASSLEVEL': {'classlevel': [0, 2]},
            'CLASS': {'class': [0, 3]},
            # ut1 (ATA1)
            # [HA in Class, HG in Class, WM, WM Grade, TScore, X, X]
            # [0, 6, 21, 28, 43, 51, 66, 81]
            'SCORE_ATA1': {'ut1': [21, 28]},
            # OM_ATA1	OMBYCLASS + OMBYCLASSLEVEL + OMBYSTREAM + OMBYSUBJGRP + OMCrossClass
            'OM_ATA1':  {'ut1_ClassRank': [0, 10], 'ut1_FormRank': [10, 20]},
            # [獎勵分, 優點, 小功, 大功, X, 缺點, 小過, 大過, 警告, X]
            # [0, 6, 12, 18, 24, 30, 36, 42, 48, 54, 60],
            # 'ANP_ATA1': {'獎勵分1': [0, 6], '優點1': [6, 12], '小功1': [12, 18], '大功1': [18, 24],
            #             '缺點1': [30, 36], '小過1': [36, 42], '大過1': [42, 48], '警告1': [48, 54]},
            # [Absent, EarlyLeave, Late]
            # [6, 12, 18],
            # 'ATT_ATA1': {'Absent1': [0, 6], 'EarlyLeave1': [6, 12], 'Late1': [12, 18]},
            # exam1 (ATA2)  = total1 (ATA3)
            # 'SCORE_ATA2': {'exam1': [21, 28]},
            # 'OM_ATA2': {'exam1_ClassRank': [0, 10], 'exam1_FormRank': [10, 20]},
            # total1 (ATA3)
            'SCORE_ATA3': {'wm1': [21, 28]},
            'OM_ATA3': {'wm1_ClassRank': [0, 10], 'wm1_FormRank': [10, 20]},
            # [獎勵分, 優點, 小功, 大功, X, 缺點, 小過, 大過, 警告, X]
            # [0, 6, 12, 18, 24, 30, 36, 42, 48, 54, 60],
            'ANP_ATA3': {'獎勵分1': [0, 6], '優點1': [6, 12], '小功1': [12, 18], '大功1': [18, 24],
                         '缺點1': [30, 36], '小過1': [36, 42], '大過1': [42, 48], '警告1': [48, 54]},
            # [Absent, EarlyLeave, Late]
            # [6, 12, 18],
            'ATT_ATA3': {'Absent1': [0, 6], 'EarlyLeave1': [6, 12], 'Late1': [12, 18]},
            # ut2
            'SCORE_ATA4': {'ut2': [21, 28]},
            'OM_ATA4': {'ut2_ClassRank': [0, 10], 'ut2_FormRank': [10, 20]},
            # exam2
            # 'SCORE_ATA5': {'wm2': [21, 28]},
            # 'OM_ATA5': {'wm2_ClassRank': [0, 10], 'wm2_FormRank': [10, 20]},
            # total2
            'SCORE_ATA6': {'wm2': [21, 28]},
            'OM_ATA6': {'wm2_ClassRank': [0, 10], 'wm2_FormRank': [10, 20]},
            'ANP_ATA6': {'獎勵分2': [0, 6], '優點2': [6, 12], '小功2': [12, 18], '大功2': [18, 24],
                         '缺點2': [30, 36], '小過2': [36, 42], '大過2': [42, 48], '警告2': [48, 54]},
            'ATT_ATA6': {'Absent2': [0, 6], 'EarlyLeave2': [6, 12], 'Late2': [12, 18]},
            # final
            'SCORE_ATA7': {'final': [21, 28]},
            'OM_ATA7': {'final_ClassRank': [0, 10], 'final_FormRank': [10, 20]},
            # 'ANP_ATA7': {'獎勵分2': [0, 6], '優點2': [6, 12], '小功2': [12, 18], '大功2': [18, 24],
            #              '缺點2': [30, 36], '小過2': [36, 42], '大過2': [42, 48], '警告2': [48, 54]},
            # 'ATT_ATA7': {'Absent2': [0, 6], 'EarlyLeave2': [6, 12], 'Late2': [12, 18]},
            # final
            # 'SCORE_ATA8': {'final': [21, 28]},
            # 'OM_ATA8': {'final_ClassRank': [0, 10], 'final_FormRank': [10, 20]},
            # 'ANP_ATA8': {'獎勵分8': [0, 6], '優點8': [6, 12], '小功8': [12, 18], '大功8': [18, 24],
            #              '缺點8': [30, 36], '小過8': [36, 42], '大過8': [42, 48], '警告8': [48, 54]},
            # 'ATT_ATA8': {'Absent8': [0, 6], 'EarlyLeave8': [6, 12], 'Late8': [12, 18]},
            }

        self.column_view = [
            'SCHYEAR', 'REGNO', 'ENNAME', 'CHNAME', 'SEX',
            'HKID', 'DOB', 'classlevel', 'class', 'CLASSNO',
            'ut1', 'ut1_ClassRank', 'ut1_FormRank',
            'wm1', 'wm1_ClassRank', 'wm1_FormRank',
            'ut2', 'ut2_ClassRank', 'ut2_FormRank',
            'wm2', 'wm2_ClassRank', 'wm2_FormRank',
            'final', 'final_ClassRank', 'final_FormRank',
            '獎勵分1', '優點1', '小功1', '大功1',
            '警告1', '缺點1',  '小過1', '大過1',
            'Absent1', 'EarlyLeave1', 'Late1',
            '獎勵分2', '優點2', '小功2', '大功2',
            '警告2', '缺點2', '小過2', '大過2',
            'Absent2', 'EarlyLeave2', 'Late2']
        self.df = pd.DataFrame(columns=self.column)
        self.view_df = None

    def view(self):
        print('create a view of score for later')
        self.view_df = self.df.copy()
        self.view_df['SCHYEAR'] = self.view_df['SCHYEAR'].apply(pd.to_numeric, errors='ignore')
        self.view_df['CLASSNO'] = self.view_df['CLASSNO'].apply(pd.to_numeric, errors='ignore')

        om_list_filter = ['OM_ATA1', 'OM_ATA2', 'OM_ATA3', 'OM_ATA4', 'OM_ATA5', 'OM_ATA6', 'OM_ATA7', 'OM_ATA8']
        for column, split_dict in self.column_to_view.items():
            for key, item in split_dict.items():
                if column in om_list_filter:
                    self.view_df['x'] = self.view_df[column].str[item[0]:item[1]]
                    self.view_df[key] = self.view_df['x'].str.split('/').str[0].apply(
                        pd.to_numeric, errors='ignore')
                else:
                    self.view_df[key] = self.view_df[column].str[item[0]:item[1]].apply(
                        pd.to_numeric, errors='ignore')
        self.view_df.to_excel('test10.xlsx')
        print(self.view_df.columns)
        # self.view_df = self.view_df[self.column_view]
        return self.view_df


# reportP conduct file setting
class ReportConduct:
    def __init__(self):
        self.name = 'conduct'
        self.filename = 'CONDUCT.xls'
        self.column = [
            'REGNO', 'ENCONDUCTNAME', 'CHCONDUCTNAME', 'DISPLAYORDER',
            'CONDUCTGRADE_ATA1', 'ENCONDUCTGRADE_ATA1', 'CHCONDUCTGRADE_ATA1',
            'CONDUCTGRADE_ATA2', 'ENCONDUCTGRADE_ATA2', 'CHCONDUCTGRADE_ATA2',
            'CONDUCTGRADE_ATA3', 'ENCONDUCTGRADE_ATA3', 'CHCONDUCTGRADE_ATA3',
            'CONDUCTGRADE_ATA4', 'ENCONDUCTGRADE_ATA4', 'CHCONDUCTGRADE_ATA4',
            'CONDUCTGRADE_ATA5', 'ENCONDUCTGRADE_ATA5', 'CHCONDUCTGRADE_ATA5',
            'CONDUCTGRADE_ATA6', 'ENCONDUCTGRADE_ATA6', 'CHCONDUCTGRADE_ATA6',
            'CONDUCTGRADE_ATA7', 'ENCONDUCTGRADE_ATA7', 'CHCONDUCTGRADE_ATA7',
            'CONDUCTGRADE_ATA8', 'ENCONDUCTGRADE_ATA8', 'CHCONDUCTGRADE_ATA8',
            ]
        self.column_filter = [
            'REGNO', 'ENCONDUCTNAME', 'CHCONDUCTNAME',
            # 'CONDUCTGRADE_ATA1',    # ut1
            # 'CONDUCTGRADE_ATA2',    # exam1
            'CONDUCTGRADE_ATA3',    # total1
            # 'CONDUCTGRADE_ATA4',    # ut2
            # 'CONDUCTGRADE_ATA5',    # exam2
            # 'CONDUCTGRADE_ATA6',    # total2
            'CONDUCTGRADE_ATA7',    # final
            # 'CONDUCTGRADE_ATA8',    # dummy
            ]
        self.column_view = self.column_filter
        self.conduct_item_list = {'守時', '儀容', '呈交功課表現', '禮貌', '責任感'}
        self.conduct_term = {1: {'item_list': ['守時', '儀容', '呈交功課表現'],
                                 'column': 'CONDUCTGRADE_ATA3'},
                             2: {'item_list': ['守時', '儀容', '呈交功課表現', '禮貌', '責任感'],
                                 'column': 'CONDUCTGRADE_ATA7'}}
        self.df = pd.DataFrame(columns=self.column)


class ReportScore:
    def __init__(self):
        self.name = 'score'
        self.filename = 'SCORE.xls'
        self.column = [
            'REGNO', 'SUBJCODE', 'SUBJCOMCODE', 'MOI',
            'CROSSCLSGRPDESC', 'CROSSCLSGRP', 'CROSSCLSSUBGRP',
            'SUBJGRPDESC', 'SUBJGRPCODE', 'SUBJTYPE',
            'ENSUBJNAME', 'CHSUBJNAME',
            'ENSUBJCOMNAME', 'CHSUBJCOMNAME',
            'ENMOI', 'CHMOI', 'PRINTSEQ',
            'EXTRA1', 'EXTRA2', 'EXTRA3', 'EXTRA4', 'EXTRA5',
            'EXTRA6', 'EXTRA7', 'EXTRA8', 'EXTRA9', 'EXTRA10',
            'EXTRA11', 'EXTRA12', 'EXTRA13', 'EXTRA14', 'EXTRA15',
            'EXTRA16', 'EXTRA17', 'EXTRA18', 'EXTRA19', 'EXTRA20',
            'SETTING_ATA1', 'SCORE_ATA1', 'GRADECODE_ATA1', 'ENGRADE_ATA1',
            'CHGRADE_ATA1', 'OM_ATA1', 'ENCOMMENT_ATA1', 'CHCOMMENT_ATA1',
            'SETTING_ATA2', 'SCORE_ATA2', 'GRADECODE_ATA2', 'ENGRADE_ATA2',
            'CHGRADE_ATA2', 'OM_ATA2', 'ENCOMMENT_ATA2', 'CHCOMMENT_ATA2',
            'SETTING_ATA3', 'SCORE_ATA3', 'GRADECODE_ATA3', 'ENGRADE_ATA3',
            'CHGRADE_ATA3', 'OM_ATA3', 'ENCOMMENT_ATA3', 'CHCOMMENT_ATA3',
            'SETTING_ATA4', 'SCORE_ATA4', 'GRADECODE_ATA4', 'ENGRADE_ATA4',
            'CHGRADE_ATA4', 'OM_ATA4', 'ENCOMMENT_ATA4', 'CHCOMMENT_ATA4',
            'SETTING_ATA5', 'SCORE_ATA5', 'GRADECODE_ATA5', 'ENGRADE_ATA5',
            'CHGRADE_ATA5', 'OM_ATA5', 'ENCOMMENT_ATA5', 'CHCOMMENT_ATA5',
            'SETTING_ATA6', 'SCORE_ATA6', 'GRADECODE_ATA6', 'ENGRADE_ATA6',
            'CHGRADE_ATA6', 'OM_ATA6', 'ENCOMMENT_ATA6', 'CHCOMMENT_ATA6',
            'SETTING_ATA7', 'SCORE_ATA7', 'GRADECODE_ATA7', 'ENGRADE_ATA7',
            'CHGRADE_ATA7', 'OM_ATA7', 'ENCOMMENT_ATA7', 'CHCOMMENT_ATA7',
            'SETTING_ATA8', 'SCORE_ATA8', 'GRADECODE_ATA8', 'ENGRADE_ATA8',
            'CHGRADE_ATA8', 'OM_ATA8', 'ENCOMMENT_ATA8', 'CHCOMMENT_ATA8',
            ]
        self.column_filter = [
            'REGNO',  # 'subject_key',
            'SUBJCODE', 'SUBJCOMCODE',
            # 'CROSSCLSGRPDESC',
            'CROSSCLSGRP', 'CROSSCLSSUBGRP',
            # 'SUBJGRPDESC', 'SUBJGRPCODE', 'SUBJTYPE',
            'ENSUBJNAME',
            # 'CHSUBJNAME',
            'ENSUBJCOMNAME',
            # 'CHSUBJCOMNAME',
            # 'ENMOI', 'CHMOI', 'PRINTSEQ',
            'SCORE_ATA1', 'OM_ATA1',                                # ut1
            'SCORE_ATA2', 'OM_ATA2',                                # exam1
            'SCORE_ATA3', 'OM_ATA3',                                # total1
            'SCORE_ATA4', 'OM_ATA4',                                # ut2
            'SCORE_ATA5', 'OM_ATA5',                                # exam2
            'SCORE_ATA6', 'OM_ATA6',                                # total2
            'SCORE_ATA7', 'OM_ATA7',                                # dummy
            'SCORE_ATA8', 'OM_ATA8',                                # final
            ]

        self.column_ata = [
            'SCORE_ATA1',   # ut1
            'SCORE_ATA2',   # exam1
            'SCORE_ATA3',   # total1, mock
            'SCORE_ATA4',   # ut2
            'SCORE_ATA5',   # exam2
            'SCORE_ATA6',   # total2
            'SCORE_ATA7',   # dummy
            'SCORE_ATA8',   # final
            ]

        # OM_ATA1	OMBYCLASS + OMBYCLASSLEVEL + OMBYSTREAM + OMBYSUBJGRP + OMCrossClass
        self.om_ata_to_rank = {
            'OM_ATA1': {'ClassRank1': [0, 10], 'FormRank1': [10, 20], 'GroupRank1': [40, 50]},  # ut1
            # 'OM_ATA2': {'ClassRank2': [0, 10], 'FormRank2': [10, 20], 'GroupRank2': [40, 50]},  # exam1
            'OM_ATA3': {'ClassRank3': [0, 10], 'FormRank3': [10, 20], 'GroupRank3': [40, 50]},  # total1
            'OM_ATA4': {'ClassRank4': [0, 10], 'FormRank4': [10, 20], 'GroupRank4': [40, 50]},  # ut2
            # 'OM_ATA5': {'ClassRank5': [0, 10], 'FormRank5': [10, 20], 'GroupRank5': [40, 50]},  # exam2
            'OM_ATA6': {'ClassRank6': [0, 10], 'FormRank6': [10, 20], 'GroupRank6': [40, 50]},  # total2
            'OM_ATA7': {'ClassRank7': [0, 10], 'FormRank7': [10, 20], 'GroupRank7': [40, 50]},  # final
            'OM_ATA8': {'ClassRank8': [0, 10], 'FormRank8': [10, 20], 'GroupRank8': [40, 50]},  # dummy
            }
        self.column_view = [
            'REGNO', 'subject_key',
            'SUBJCODE', 'SUBJCOMCODE',
            'CROSSCLSGRP', 'CROSSCLSSUBGRP',
            'ENSUBJNAME', 'ENSUBJCOMNAME',
            'SCORE_ATA1', 'ClassRank1', 'FormRank1', 'GroupRank1',  # ut1
            # 'SCORE_ATA2', 'ClassRank2', 'FormRank2', 'GroupRank2',  # ut1
            'SCORE_ATA3', 'ClassRank3', 'FormRank3', 'GroupRank3',  # ut1
            'SCORE_ATA4', 'ClassRank4', 'FormRank4', 'GroupRank4',  # ut1
            # 'SCORE_ATA5', 'ClassRank5', 'FormRank5', 'GroupRank5',  # ut1
            'SCORE_ATA6', 'ClassRank6', 'FormRank6', 'GroupRank6',  # ut1
            'SCORE_ATA7', 'ClassRank7', 'FormRank7', 'GroupRank7',  # ut1
            'SCORE_ATA8', 'ClassRank8', 'FormRank8', 'GroupRank8',  # ut1
            ]
        self.column_numeric = [
            'SCORE_ATA1', 'ClassRank1', 'FormRank1', 'GroupRank1',  # ut1
            #        'SCORE_ATA2', 'ClassRank2', 'FormRank2', 'GroupRank2',  # ut1
            'SCORE_ATA3', 'ClassRank3', 'FormRank3', 'GroupRank3',  # exam1
            'SCORE_ATA4', 'ClassRank4', 'FormRank4', 'GroupRank4',  # ut2
            #        'SCORE_ATA5', 'ClassRank5', 'FormRank5', 'GroupRank5',  # ut1
            'SCORE_ATA6', 'ClassRank6', 'FormRank6', 'GroupRank6',  # exam2
            'SCORE_ATA7', 'ClassRank7', 'FormRank7', 'GroupRank7',  # final
            'SCORE_ATA8', 'ClassRank8', 'FormRank8', 'GroupRank8',  # final
            ]
        self.df = pd.DataFrame(columns=self.column)
        self.view_df = None

    def view(self):
        # split rankings
        # convert numbers to numbers
        subject_name = {'Chinese History': 'chs',
                        'Chinese Language': 'chi',
                        'Computer Literacy': 'cps',
                        'Design And Technology': 'dte',
                        'English Language': 'eng',
                        'Ethics/ Religious Education': 'bik',
                        'Geography': 'geo',
                        'Home Economics': 'hec',
                        'Science(Secondary 1-3)': 'isc',
                        'Liberal Studies': 'lst',
                        'Mathematics': 'mth',
                        'Music': 'mus',
                        'Physical Education Lessons': 'ped',
                        'Putonghua': 'pth',
                        'Visual Arts': 'via',
                        'History': 'hst',
                        'Biology': 'bio',
                        'Chemistry': 'chm',
                        'Economics': 'eco',
                        'Physics': 'phy',
                        'BAFS (Accounting)': 'baf',
                        'Mathematics (Compulsory Part)': 'mth',
                        'Mathematics (Extended Part – Module 1)': 'm1',
                        'Mathematics (Extended Part – Module 2)': 'm2',
                        'Information & Communication Technology': 'ict',
                        'STEM': 'stm',
                        }

        print('create a view of score for later')
        print('\tadd rankings in score.df')

        self.view_df = self.df.copy()

        # om_ata to class_rank / form_rank / group_rank
        for column, rank_dict in self.om_ata_to_rank.items():
            for key, item in rank_dict.items():
                self.view_df['x'] = self.view_df[column].str[item[0]:item[1]]
                self.view_df[key] = (self.view_df['x'].str.split('/').str[0]
                                         .apply(pd.to_numeric, errors='ignore'))

        # convert all ata from str to numeric
        for column in self.column_ata:
            self.view_df[column] = self.view_df[column].apply(pd.to_numeric, errors='ignore')

        # create a subject key containing component code
        self.view_df['subject'] = self.view_df['ENSUBJNAME'].apply(lambda x: subject_name[x])
        self.view_df['subject_key'] = (np.where(self.view_df['SUBJCOMCODE'].str.len() > 0,
                                                'x' + self.view_df['SUBJCODE'] + self.view_df['SUBJCOMCODE'],
                                                'x' + self.view_df['SUBJCODE']))
        self.view_df = self.view_df[self.column_view]
        self.view_df.to_excel('score_view.xlsx')
        return self.view_df
        # merge conduct into student


subject_from_code = {
    'x080': 'chi', 'x08001': 'chi01', 'x08002': 'chi02', 'x08003': 'chi03', 'x08004': 'chi04', 'x08005': 'chi05',
    'x165': 'eng', 'x16501': 'eng01', 'x16502': 'eng02', 'x16503': 'eng03', 'x16504': 'eng04',
    'x280': 'mth', 'x22S': 'mth', 'x23S': 'm1', 'x24S': 'm2',
    'x075': 'chs', 'x235': 'hst', 'x210': 'geo', 'x135': 'eco',
    'x260': 'isc', 'x315': 'phy', 'x070': 'chm', 'x045': 'bio',
    'x12N': 'baf', 'x81N': 'ict',
    'x265': 'lst',
    'x185': 'bik',
    'x110': 'cps', 'x350': 'pth', 'x432': 'via', 'x130': 'dte', 'x240': 'hec', 'x909': 'stm',
    'x300': 'mus', 'x310': 'ped',
    }

#   rpt_n.py
#   (0) copy analysis statistics
#   (1) unzip all reportP zip files
#   (2) combine all student, conduct, score into a df
#   (3) write df into rptP-exam template
#   (4) from df build view and ranks files

# loop to unzip junior reportP
# need to check which classlevel need to combine
#
#   exam_type: ut1, exam1, ut2, mock, final   (no exam2)
#


class Report:
    def __init__(self, exam_year: str, exam_type: str):
        self.exam_year = exam_year
        self.exam_type = exam_type
        self.home_folder = os.path.abspath(os.getcwd())
        self.report_home = os.path.join(self.home_folder, self.exam_year + 'report')
        self.report_zip_folder = os.path.join(self.report_home, 'report_zip')
        self.report_unzip_folder = os.path.join(self.report_home, 'unzip')
        # move template back inside Report Home
        self.report_template_folder = os.path.join(self.report_home, 'template', 'excel')
        self.report_exam_folder = os.path.join(self.report_home, 'report_' + self.exam_type)

        if self.exam_type in ['exam1', 'ut2', 'final']:
            # self.classlevel_list = ['s1', 's2', 's3', 's4', 's5']
            self.classlevel_list = ['s1', 's2', 's3', 's4', 's5']
        elif self.exam_type == 'mock':
            self.classlevel_list = ['s6']
        else:
            self.classlevel_list = ['s1', 's2', 's3', 's4', 's5', 's6']

    # file setup
        self.exam_run_file = os.path.join(self.home_folder, self.exam_year + '-exam-run.xlsm')
        self.copy_sheet = ['sick', 'dsewm', 'statistics', 'promotion', 'award']

        self.report_file_dict = {'student': 'STUDENT.xls', 'score': 'SCORE.xls', 'conduct': 'CONDUCT.xls'}
        self.report_df_dict = {'conduct': ReportConduct(),
                               'student': ReportStudent(),
                               'score': ReportScore()}

        self.report_template = os.path.join(self.report_template_folder, 'report-template.xlsx')
        self.report_template_empty = os.path.join(self.report_template_folder, 'report-template-empty.xlsx')
        self.report_all_file = os.path.join(self.report_exam_folder,
                                            self.exam_year + '-report-' + self.exam_type + '-all.xlsx')
        self.rank_template = os.path.join(self.report_template_folder, 'template.xlsx')
        self.rank_file = os.path.join(self.report_home, self.exam_year + '-' + self.exam_type + '-awardee.xlsx')

        self.view_df = None
        self.view_file = os.path.join(self.report_home, self.exam_year + '-' + self.exam_type + '-report-view.xlsx')

        self.view_column = [
            'REGNO', 'HKID', 'STRN', 'CHNAME', 'ENNAME', 'SEX', 'DOB', 'SCHYEAR', 'CLASSNO',
            'PROMOTESTATUS', 'STAFFCODE', 'CHSTAFFNAME',
            'SCORE_ATA1', 'OM_ATA1', 'ANP_ATA1', 'ATT_ATA1', 'SCORE_ATA2', 'OM_ATA2', 'ANP_ATA2', 'ATT_ATA2',
            'SCORE_ATA3', 'OM_ATA3', 'ANP_ATA3', 'ATT_ATA3', 'SCORE_ATA4', 'OM_ATA4', 'ANP_ATA4', 'ATT_ATA4',
            'SCORE_ATA5', 'OM_ATA5', 'ENCOMMENT_ATA5', 'CHCOMMENT_ATA5', 'ANP_ATA5', 'ATT_ATA5',
            'SCORE_ATA6', 'OM_ATA6', 'ENCOMMENT_ATA6', 'CHCOMMENT_ATA6', 'ANP_ATA6', 'ATT_ATA6',
            'SCORE_ATA7', 'OM_ATA7', 'ENCOMMENT_ATA7', 'CHCOMMENT_ATA7', 'ANP_ATA7', 'ATT_ATA7',
            'SCORE_ATA8', 'OM_ATA8', 'ENCOMMENT_ATA8', 'CHCOMMENT_ATA8', 'ANP_ATA8', 'ATT_ATA8',
            'classlevel', 'class', 'ut1', 'x', 'ut1_ClassRank', 'ut1_FormRank', 'wm1', 'wm1_ClassRank', 'wm1_FormRank',
            '獎勵分1', '優點1', '小功1', '大功1', '缺點1',
            '小過1', '大過1', '警告1', 'Absent1', 'EarlyLeave1', 'Late1',
            'ut2', 'ut2_ClassRank', 'ut2_FormRank', 'wm2', 'wm2_ClassRank', 'wm2_FormRank',
            '獎勵分2', '優點2', '小功2', '大功2', '缺點2', '小過2', '大過2', '警告2', 'Absent2', 'EarlyLeave2', 'Late2',
            'final', 'final_ClassRank', 'final_FormRank', 'CHCONDUCTNAME', 'CONDUCTGRADE_ATA7',
            'chi', 'chi_rank', 'eng', 'eng_rank',
            'mth_x', 'mth_rank_x', 'mth_y', 'mth_rank_y', 'm1', 'm1_rank', 'm2', 'm2_rank',
            'chs', 'chs_rank', 'hst', 'hst_rank', 'geo', 'geo_rank', 'eco', 'eco_rank',
            'isc', 'isc_rank', 'phy', 'phy_rank', 'chm', 'chm_rank', 'bio', 'bio_rank',
            'baf', 'baf_rank', 'ict', 'ict_rank', 'lst', 'lst_rank',
            'bik', 'bik_rank', 'cps', 'cps_rank', 'pth', 'pth_rank', 'via', 'via_rank',
            'stm', 'stm_rank', 'dte', 'dte_rank', 'hec', 'hec_rank', 'mus', 'mus_rank', 'ped', 'ped_rank',
            ]

        self.wm_rank_col_common = ['REGNO', 'ENNAME', 'CHNAME', 'SEX', 'classlevel', 'class', 'CLASSNO']

        if self.exam_type == 'ut1':
            self.wm_class_rank_col = ['ut1', 'ut1_ClassRank']
            self.wm_form_rank_col = ['ut1', 'ut1_FormRank']
            self.class_rank_col = ['SCORE_ATA1', 'ClassRank1']
            self.form_rank_col = ['SCORE_ATA1', 'FormRank1']
            self.group_rank_col = ['SCORE_ATA1', 'GroupRank1']

        elif self.exam_type == 'exam1':
            self.wm_class_rank_col = ['wm1', 'wm1_ClassRank']
            self.wm_form_rank_col = ['wm1', 'wm1_FormRank']
            self.class_rank_col = ['SCORE_ATA3', 'ClassRank3']
            self.form_rank_col = ['SCORE_ATA3', 'FormRank3']
            self.group_rank_col = ['SCORE_ATA3', 'GroupRank3']

        elif self.exam_type == 'ut2':
            self.wm_class_rank_col = ['ut2', 'ut2_ClassRank']
            self.wm_form_rank_col = ['ut2', 'ut2_FormRank']
            self.class_rank_col = ['SCORE_ATA4', 'ClassRank4']
            self.form_rank_col = ['SCORE_ATA4', 'FormRank4']
            self.group_rank_col = ['SCORE_ATA4', 'GroupRank4']

        elif self.exam_type == 'exam2':
            self.wm_class_rank_col = ['wm2', 'wm2_ClassRank']
            self.wm_form_rank_col = ['wm2', 'wm2_FormRank']
            self.class_rank_col = ['SCORE_ATA6', 'ClassRank6']
            self.form_rank_col = ['SCORE_ATA6', 'FormRank6']
            self.group_rank_col = ['SCORE_ATA6', 'GroupRank6']

        elif self.exam_type == 'final':
            self.wm_class_rank_col = ['final', 'final_ClassRank']
            self.wm_form_rank_col = ['final', 'final_FormRank']
            self.class_rank_col = ['SCORE_ATA7', 'ClassRank7']
            # self.class_rank_col = ['SCORE_ATA8', 'ClassRank8']
            self.form_rank_col = ['SCORE_ATA7', 'FormRank7']  # final
            # self.form_rank_col = ['SCORE_ATA8', 'FormRank8']
            self.group_rank_col = ['SCORE_ATA7', 'GroupRank7']  # final
            # self.group_rank_col = ['SCORE_ATA8', 'GroupRank8']
            print(type(self.form_rank_col))

        elif self.exam_type == 'mock':
            self.wm_class_rank_col = ['final', 'final_ClassRank']
            self.wm_form_rank_col = ['final', 'final_FormRank']
            self.class_rank_col = ['SCORE_ATA3', 'ClassRank3']
            self.form_rank_col = ['SCORE_ATA3', 'FormRank3']
            self.group_rank_col = ['SCORE_ATA3', 'FormRank3']
        # awardee file setting
        # use template to save the follow 2 files

    def unzip(self):
        # if not classlevel_list:
        #    classlevel_list = self.classlevel_list
        # print('class level:', classlevel_list)
        print('extracting report zip files')

        # zip file : REPORTP_exam1_s1.zip
        # if only 1 file to product,
        # then no need to choose classlevel
        with os.scandir(self.report_zip_folder) as report_zip_files:
            for entry in report_zip_files:
                exam_type = entry.name.split('_')[1]
                classlevel = entry.name.split('_')[2][:2]

                # zip_filename = os.path.join(self.report_zip_folder, 'REPORTP_' + self.exam_type + classlevel + '.zip')
                if entry.name[-3:] == 'zip':
                    extract_folder = os.path.join(self.report_unzip_folder, self.exam_type, classlevel)
                    if not os.path.exists(extract_folder):
                        try:
                            print(f'\tmkdir { extract_folder }')
                            os.mkdir(extract_folder)
                        except OSError:
                            print(f'\terror: { extract_folder } is not ready.')

                    with zipfile.ZipFile(file=entry, mode='r') as z:
                        # zip.printdir()
                        print(f'extracting { classlevel } { exam_type }')
                        for key, file in self.report_file_dict.items():
                            print(f'\t{ classlevel } { file } ...')
                            z.extract(file, path=extract_folder)

    def load_file_to_df(self):
        # load through 1 type of file of all classlevel
        # and then create their df
        # and then merge all of them
        print('merge student, conduct and score files')
        temp_classlevel_list = self.classlevel_list
        for key, item in self.report_df_dict.items():
            print(f'\t{ key } { item }')
            for classlevel in self.classlevel_list:
                filename = self.report_file_dict[key]
                file_to_load = os.path.join(self.report_unzip_folder, self.exam_type, classlevel, filename)
                if os.path.isfile(file_to_load):
                    temp_column = self.report_df_dict[key].column
                    temp_df = pd.read_excel(file_to_load,
                                            sheet_name=key.upper(),
                                            converters=converter(temp_column))
                    item.df = item.df.append(temp_df[temp_column])
                else:
                    temp_classlevel_list.remove(classlevel)
                    print(f'\t{ classlevel }: { file_to_load } does not exist')
                # print('{}: row {} col {}'.format(classlevel, len(item.df), len(item.df.columns)))

            item.df[item.df.columns].reset_index(inplace=True)
            print(f'classlevel: {temp_classlevel_list }')
            print(f'{ key }: row { len(item.df) } col { len(item.df.columns) }')

        for key in ['student', 'score']:
            print(f'\tcreate { key } view')
            view_file = os.path.join(self.report_exam_folder,
                                     self.exam_year + '-' + self.exam_type + '-' + key + '-view.xlsx')
            self.report_df_dict[key].view()
            self.report_df_dict[key].view_df.to_excel(view_file, index=False)

    def write_to_report(self):
        print('\nwrite student, conduct, score to rpt template')
        # for key in ['student', 'score', 'conduct']:
        #     self.report_df_dict[key].df.to_excel(key + '.xlsx')

        temp_df = self.report_df_dict['student'].df
        report_wb = openpyxl.load_workbook(filename=self.report_template_empty, data_only=False)
        temp_ws = report_wb['STUDENT']

        temp_col = self.report_df_dict['student'].column
        # print('temp_col: {}'.format(temp_col))
        temp_col_dict = dict(zip(temp_col, range(1, len(temp_col)+1)))
        print('\twriting {} sheet.'.format('student'))
        cur_row = 2
        # for record in temp_df.sort_values(
        #             by=['CLASSLEVEL', 'CLASS', 'CLASSNO'],
        #             ascending=[True, True, True]).to_dict(orient='records'):
        for record in temp_df.to_dict(orient='records'):
            for key, item in record.items():
                if key in temp_col:
                    if item is not None:
                        temp_ws.cell(row=cur_row, column=temp_col_dict[key]).value = item
                    else:
                        pass
                elif key == 'index':
                    pass
                else:
                    print(f'key { key } not found')
            cur_row += 1

        # write conductDF into conduct sheet
        temp_ws = report_wb['CONDUCT']
        temp_df = self.report_df_dict['conduct'].df
        temp_col = self.report_df_dict['conduct'].column
        temp_col_dict = dict(zip(temp_col, range(1, len(temp_col)+1)))
        print('\twriting {} sheet.'.format('conduct'))
        cur_row = 2
        for record in temp_df.to_dict(orient='records'):
            for key, item in record.items():
                if key in temp_col:
                    if item is not None:
                        temp_ws.cell(row=cur_row, column=temp_col_dict[key]).value = item
                    else:
                        pass
                elif key == 'index':
                    pass
                else:
                    print(f'key { key } not found')
            cur_row += 1

        # write conductDF into conduct sheet
        temp_ws = report_wb['SCORE']
        temp_df = self.report_df_dict['score'].df
        temp_col = self.report_df_dict['score'].column
        temp_col_dict = dict(zip(temp_col, range(1, len(temp_col)+1)))
        print('\twriting {} sheet.'.format('score'))
        cur_row = 2
        for record in temp_df.to_dict(orient='records'):
            # print(temp_df.head(5).to_dict(orient='records'))
            # input()
            for key, item in record.items():
                # print(key)
                if key in temp_col:
                    if item is not None:
                        temp_ws.cell(row=cur_row, column=temp_col_dict[key]).value = item
                    else:
                        pass
                elif key == 'index':
                    pass
                else:
                    print(f'key { key } not found')
            cur_row += 1

        report_wb.save(self.report_all_file)
        report_wb.close()
        print(f'report file generated: { self.report_all_file }')

    def view(self):
        # merge conduct into student
        # use openpyxl to load a template
        # set up view file and rank file

        master_subject_view_list = {
            'x080': ['chi', 'Chinese Language'],
            'x165': ['eng', 'English Language'],
            'x280': ['mth', 'Mathematics'],
            'x22S': ['mth', 'Mathematics (Compulsory Part)'],
            'x23S': ['m1', 'Mathematics (Extended Part – Module 1)'],
            'x24S': ['m2', 'Mathematics (Extended Part – Module 2)'],
            'x075': ['chs', 'Chinese History'],
            'x235': ['hst', 'History'],
            'x210': ['geo', 'Geography'],
            'x135': ['eco', 'Economics'],
            'x260': ['isc', 'Science(Secondary 1-3)'],
            'x315': ['phy', 'Physics'],
            'x070': ['chm', 'Chemistry'],
            'x045': ['bio', 'Biology'],
            'x12N': ['baf', 'BAFS (Accounting)'],
            'x81N': ['ict', 'Information & Communication Technology'],
            'x265': ['lst', 'Liberal Studies'],
            'x185': ['bik', 'Ethics/ Religious Education'],
            'x110': ['cps', 'Computer Literacy'],
            'x350': ['pth', 'Putonghua'],
            'x432': ['via', 'Visual Arts'],
            'x130': ['dte', 'Design And Technology'],
            'x240': ['hec', 'Home Economics'],
            'x300': ['mus', 'Music'],
            'x310': ['ped', 'Physical Education Lessons'],
            'x909': ['stm', 'STEM']
        }

        view_wb = openpyxl.load_workbook(self.rank_template)
        view_writer = pd.ExcelWriter(self.view_file, engine='openpyxl')
        view_writer.book = view_wb

        self.view_df = self.report_df_dict['student'].view_df.copy()
#        self.conduct_term = {1: {'item_list': ['守時', '儀容', '呈交功課表現'],
#                                 'column': 'CONDUCTGRADE_ATA3'},
#                             2: {'item_list': ['守時', '儀容', '呈交功課表現', '禮貌', '責任感'],
#                                 'column': 'CONDUCTGRADE_ATA7'}}

        print('\tmerge student and conduct to master view')
        for key, conduct in self.report_df_dict['conduct'].conduct_term.items():
            # loop through conduct items and merge to student df immediately
            print('\tmerge {}: {}'.format(key, conduct['column']))
            temp_conduct_df = self.report_df_dict['conduct'].df[['REGNO', 'CHCONDUCTNAME', conduct['column']]]

            for item in conduct['item_list']:
                k = 1
                # print('\titem: {}'.format(item))
                # temp_df = temp_conduct_df[temp_conduct_df['CHCONDUCTNAME'] == item]
                self.view_df = pd.merge(self.view_df,
                                        temp_conduct_df[temp_conduct_df['CHCONDUCTNAME'] == item],
                                        how='outer', on='REGNO')
                self.view_df = self.view_df.rename(columns={conduct['column']: item + str(key)})

        print(self.form_rank_col)
        temp_subjects_filter = ['REGNO', 'subject_key'] + self.form_rank_col
        score_df = (self.report_df_dict['score'].view_df[temp_subjects_filter]
                        .sort_values(by=['REGNO'], ascending=[True]))

        for key, item in master_subject_view_list.items():
            # print(key)
            subj_rank = item[0] + '_rank'
            temp_df = score_df[score_df['subject_key'] == key]
            temp_form_rank_filter = ['REGNO'] + self.form_rank_col
            form_rank_column_rename = {self.form_rank_col[0]: item[0],
                                       self.form_rank_col[1]: subj_rank}
            # rename the column index
            temp_df = temp_df[temp_form_rank_filter].rename(columns=form_rank_column_rename)
            self.view_df = pd.merge(self.view_df, temp_df, how='left', on='REGNO')

        # self.view_df.to_excel(view_writer, sheet_name='view', index=False)
        # loop to write student, score and conduct
        # for key, item in self.report_df_dict.items():
        #     print('\twrite {} {} to view file'.format(key, item))
        #     item.df[item.column_filter].to_excel(view_writer, sheet_name=key, index=False)

        filter_col = self.wm_rank_col_common + self.wm_class_rank_col
        print('\tsheet student_full')
        # print(len(self.view_df.columns), len(self.view_df[self.view_column].columns))
        self.view_df.to_excel(view_writer, sheet_name=self.exam_type, index=False)
        # self.view_df[self.view_column].to_excel(view_writer, sheet_name='all', index=False)
        print(f'\tview file: { self.view_file }')
        view_writer.save()
        view_writer.close()

    def record_card_view(self):
        student_column_filter = ['REGNO', 'CHNAME', 'ENNAME', 'SCHYEAR', 'CLASS', 'CLASSNO',
                                 'CHSTAFFNAME', 'ANPCHDESC',
                                 'SCORE_ATA7', 'OM_ATA7',
                                 'CHCOMMENT_ATA6', 'ANP_ATA6', 'ATT_ATA6', 'SCORE_ATA8', 'OM_ATA8']
        conduct_column_filter = ['REGNO', 'CHCONDUCTNAME', 'CONDUCTGRADE_ATA7']
        score_column_filter = ['REGNO', 'SUBJCODE', 'SUBJCOMCODE',
                               'ENSUBJNAME', 'ENSUBJCOMNAME',
                               'SCORE_ATA7', 'ENGRADE_ATA7', 'OM_ATA7']
        record_card_column = ['regno', 'schyear', 'classlevel', 'classcode', 'wm', 'rank',
                              '守時', '儀容', '呈交功課表現', '禮貌', '責任感',
                              'abs', 'late',
                              'ct1', 'ct2']
        subject_dict = {'080_': 'chi', '08001': 'chi01',  '08002': 'chi02', '08003': 'chi03', '08004': 'chi04',
                        '08005': 'chi05',
                        '165_': 'eng', '16501': 'eng01', '16502': 'eng02', '16503': 'eng03', '16504': 'eng04',
                        '280_': 'mth', '22S_': 'mth', '23S_': 'm1', '24S_': 'm2',
                        '075_': 'chs', '235_': 'hst', '210_': 'geo', '135_': 'eco',
                        '260_': 'isc', '315_': 'phy', '070_': 'chm', '045_': 'bio',
                        '12N_': 'baf', '81N_': 'ict',
                        '265_': 'lst', '185_': 'bik', '300_': 'mus', '310_': 'ped', '909_': 'stm',
                        '110_': 'cps', '350_': 'pth', '432_': 'via', '130_': 'dte', '240_': 'hec',
                        }

        def ap_text(ap_string):
            # 獎勵分/優點/小功/大功/缺點/小過/大過/警告
            if ap_string == '-1':
                return '****'
            else:
                ap_list = ['獎勵分', '優點', '小功', '大功', '缺點', '小過', '大過', '警告']
                a_item = ['優點', '小功', '大功']
                p_item = ['缺點', '小過', '大過']
                ap_dict = dict(zip(ap_list, [int(x) for x in ap_string.split()]))
                k = [str(n) + text if n else '' for text, n in ap_dict.items()]
                list1 = [''.join(k[1:4]), '/',  ''.join(k[4:7])]
                list2 = [x if len(x) else '****' for x in list1]
                s = ''.join(list2)
                return s

        # AbsentDay + EarlyLeaveDay + LateDay
        # English Field: (6 + 6 + 6)
        # (Lvl1 + Lvl2 + Lvl3 + Lvl4 + Lvl5 Merit) + (Lvl1 + Lvl2 + Lvl3 + Lvl4 + Lvl5 Demerit) +
        # (CalCondMark + Awarded Conduct Mark + Deducted Conduct Mark + Conduct Mark Reason)
        # (English 6 + English 6 + English 6 + English 6 + English 6 + English 6 + English 6 +
        # English 6 + English 6 + English 6 + English 7 + English 6 + English 6 + Bilingual 60)
        # CHCONDUCTNAME: 守時/儀容/呈交功課表現/禮貌/責任感

        conduct_list = ['守時', '儀容', '呈交功課表現', '禮貌', '責任感']

        stu_df = pd.read_excel(self.report_all_file,
                               'STUDENT',
                               converters=converter(student_column_filter))[student_column_filter]
        
        conduct_df = pd.read_excel(self.report_all_file, 'CONDUCT')[conduct_column_filter]
        score_df = pd.read_excel(self.report_all_file,
                                 'SCORE',
                                 converters=converter(score_column_filter))[score_column_filter]
        score_df.to_excel('score_1.xlsx')
        score_df['x'] = score_df['SUBJCOMCODE'].fillna('_')
        score_df['class_rank'] = score_df['OM_ATA7'].str[0:10]
        score_df['form_rank'] = score_df['OM_ATA7'].str[10:20]
        score_df['score'] = score_df['SCORE_ATA7']
        score_df['subject_code'] = score_df['SUBJCODE'] + score_df['x']
        score_df['subject_code'].replace(subject_dict, inplace=True)
        score_column = ['REGNO', 'subject_code', 'SCORE_ATA7', 'ENGRADE_ATA7', 'class_rank', 'form_rank']
        score_pivot = score_df[['REGNO', 'subject_code', 'score']].pivot_table(index=['REGNO'],
                                                                               columns='subject_code',
                                                                               values='score')
     # crank_pivot = score_df[['REGNO', 'subject_code', 'class_rank']].pivot_table(index=['REGNO'],
     #                                                                              columns='subject_code',
     #                                                                              values='class_rank')
     # frank_pivot = score_df[['REGNO', 'subject_code', 'form_rank']].pivot_table(index=['REGNO'],
     #                                                                              columns='subject_code',
     #                                                                              values='form_rank')
        score_df[score_column].to_excel('score.xlsx', index=False)
        score_pivot.to_excel('score_pivot.xlsx')

        for conduct in conduct_list:
            conduct_temp_df = conduct_df[conduct_df['CHCONDUCTNAME'] == conduct]
            stu_df = pd.merge(stu_df, conduct_temp_df, how='left', on='REGNO')
            stu_df.rename(columns={'CONDUCTGRADE_ATA7': conduct}, inplace=True)
            # stu_df[conduct] = stu_df['CONDUCTGRADE_ATA7']

        stu_df = pd.merge(stu_df, score_pivot, how='left', on='REGNO')

        stu_df.rename(columns={'REGNO': 'regno',
                               'SCHYEAR': 'schyear',
                               'CHNAME': 'chname',
                               'ENNAME': 'enname',
                               'CLASSNO': 'classno',
                               'CHCOMMENT_ATA6': 'comment'}, inplace=True)

        stu_df['classlevel'] = stu_df['CLASS'].str[:2]
        stu_df['classcode'] = stu_df['CLASS'].str[:3]
        stu_df = stu_df[stu_df.regno.notnull()]
        stu_df['x'] = stu_df['classno'].astype(int).astype(str)
        stu_df['key'] = stu_df['classcode'] + stu_df['classno'].astype(int).astype(str).str.zfill(2)
        stu_df['ct1'] = stu_df['CHSTAFFNAME'].str[:6]
        stu_df['ct2'] = stu_df['CHSTAFFNAME'].str[6:12]
        stu_df['wm'] = stu_df['SCORE_ATA7'].str[20:27].astype(float)
        stu_df['rank'] = stu_df['OM_ATA7'].str[10:20]
        stu_df['abs'] = stu_df['ATT_ATA6'].str[0:6].astype(float)
        stu_df['late'] = stu_df['ATT_ATA6'].str[12:18].astype(float)
        stu_df['ap_temp'] = stu_df['ANP_ATA6'].fillna('-1')
        stu_df['ap_text'] = stu_df['ap_temp'].apply(ap_text)

        record_card_column2 = ['key', 'regno', 'enname', 'chname',
                               'schyear', 'classlevel', 'classcode', 'classno',
                               'ct1', 'ct2', 'wm', 'rank', 'comment',
                               '守時', '儀容', '呈交功課表現', '禮貌', '責任感',
                               'abs', 'late', 'ap_text'] + list(subject_dict.values())

        stu_df[record_card_column2].to_excel('temp.xlsx', index=False)
        # stu_df.to_excel('temp.xlsx', index=False)

    def filter_rank(self):
        # build a list of filter to sort out
        # #1 class rank : s1, s2, s3 only
        # #2 improved : s1, s2, s3, s4, s5 only
        # #3 form rank : s1 - s6, all subjects
        # #4 group rank : s4 - s5, x-subject

        subject_rank_col_common = ['REGNO', 'ENNAME', 'CHNAME', 'SEX', 'classlevel',
                                   'class', 'CLASSNO', 'ENSUBJNAME', 'SUBJCODE']

        chi_eng_component_list = ['x08001', 'x08002', 'x08003', 'x08004', 'x08005',
                                  'x16501', 'x16502', 'x16503', 'x16504', 'x16505']

        classlevel_list_upper = list(map(str.upper, self.classlevel_list))
        rank_wb = openpyxl.load_workbook(self.rank_template)
        rank_writer = pd.ExcelWriter(self.rank_file, engine='openpyxl')
        rank_writer.book = rank_wb

        print('filter rankings')
        print(f'award file: { self.rank_file }')

        # 1.    Get Class Rank from 1 to 5
        #       Use filter on classlevel and self.wm_class_rank_col
        print('\tsheet', 'wm_rank_class')
        stu_df = self.report_df_dict['student'].view_df
        # stu_df.to_excel('stu_df.xlsx')
        # exam_type_temp = 'final'
        temp_df = stu_df[stu_df[self.wm_class_rank_col[1]] < 6]

        temp_df = (temp_df[temp_df.classlevel.isin(classlevel_list_upper)]
                   .sort_values(by=['classlevel', 'class', self.wm_class_rank_col[0]],
                                ascending=[True, True, False]))

        filter_col = self.wm_rank_col_common + self.wm_class_rank_col
        temp_df[filter_col].to_excel(rank_writer, sheet_name='wm_rank_class', index=False)

        # 2.    Get Form Rank from 1 to 5
        #       Use filter on classlevel and self.wm_form_rank_col
        print('\tsheet', 'wm_rank_form')
        temp_df = stu_df[stu_df[self.wm_form_rank_col[1]] < 6]
        temp_df = (temp_df[temp_df.classlevel.isin(classlevel_list_upper)]
                   .sort_values(by=['classlevel', self.wm_form_rank_col[0]],
                                ascending=[True, False]))
        filter_col = self.wm_rank_col_common + self.wm_form_rank_col
        temp_df[filter_col].to_excel(rank_writer, sheet_name='wm_rank_form', index=False)

        # 3.    Get Form Rank from 1 to 5
        #       Use filter on classlevel and self.wm_form_rank_col
        if self.exam_type in ['exam2', 'final']:
            print('\tsheet', 'improved')
            print(f'classlevel_list_upper: { classlevel_list_upper }')
            filter_classlevel = stu_df.classlevel.isin(classlevel_list_upper)
            temp_df2 = stu_df[filter_classlevel]
            # print(f'stu_df: { len(temp_df2)}')
            # temp_df2.to_excel('test_stu_df.xlsx')
            # print(filter_classlevel)
#            for classlevel in ['S1', 'S2', 'S3']:
            #temp_df2['wm2_x'] = temp_df2['wm2'].astype(str).strip().ap
            temp_df2['wm2x'] = temp_df2['wm2'].apply(lambda x: x if len(str(x).strip()) > 0 else -1)
            print(len(temp_df2))
            # print(temp_df['wm2'])
            # temp_df[['wm2']].to_excel('wm.xlsx')
            temp_df2.to_excel('test_stu_df.xlsx')
            temp_df2['improved'] = temp_df2['wm2x'] - temp_df2['wm1']
            temp_df2['improved_rank'] = temp_df2.groupby('class')['improved'].rank()
            temp_df2.to_excel('test.xlsx')

            #temp_df['improved_rank_max'] = temp_df.groupby('class')['improved'].rank(method='max')
            # temp_df.to_excel(rank_writer, sheet_name='temp4', index=False)
            temp_df2 = (temp_df2[temp_df2.classlevel.isin(classlevel_list_upper)]
                       .sort_values(by=['classlevel', 'class', 'improved_rank'],
                                    ascending=[True, True, False]))
            filter_col = self.wm_rank_col_common + ['wm1', 'wm2', 'improved', 'improved_rank']
            temp_df2[filter_col].to_excel(rank_writer, sheet_name='improved_rank_class', index=False)
        else:
            print(f'\timproved list will be skipped in { self.exam_type }')

        # 4. Filter Class 1-6 of each subject
        #    Need to filter out chi and eng components
        print('\tsheet', 'subj_rank_class')

        # temporarily override self.exam_type2
        # examtype2_temp = self.exam_type2
        # examtype2 = 'exam1'

        temp_df = pd.merge(self.report_df_dict['score'].view_df[
                               self.report_df_dict['score'].view_df[self.class_rank_col[1]] < 6],
                           self.report_df_dict['student'].view_df[
                               ['REGNO', 'ENNAME', 'CHNAME', 'SEX', 'classlevel', 'class', 'CLASSNO']],
                           how='left', on='REGNO')
        temp_df = (temp_df[~temp_df.subject_key.isin(chi_eng_component_list)]
                   .sort_values(by=['classlevel', 'class', 'SUBJCODE', self.class_rank_col[1]],
                                ascending=[True, True, True, True]))
        filter_col = subject_rank_col_common + self.class_rank_col
        temp_df[filter_col].to_excel(rank_writer, sheet_name='subj_rank_class', index=False)

        # 4. Filter Class 1-6 of each subject
        #    Need to take care ped male and female
        print('sheet', 'subj_rank_form')
        temp_df = pd.merge(self.report_df_dict['score'].view_df[
                               self.report_df_dict['score'].view_df[self.form_rank_col[1]] < 15],
                           self.report_df_dict['student'].view_df[
                               ['REGNO', 'ENNAME', 'CHNAME', 'SEX', 'classlevel', 'class', 'CLASSNO']],
                           how='left', on='REGNO')
        # print(temp_df['score']['subject_key'].map(subject_from_code).tail(5))
        # temp_df['score']['subject'] = np.where(temp_df['score']['subject'] == 'ped',
        #                                            temp_df['score']['subject'] + temp_df['score']['SEX'],
        #                                            temp_df['score']['subject'])
        # temp_df.to_excel(rank_writer, sheet_name='raw12', index=False)

        # filter away chi/eng components
        temp_df = (temp_df[~temp_df.subject_key.isin(chi_eng_component_list)]
                   .sort_values(by=['classlevel', 'subject_key', self.form_rank_col[1]],
                                ascending=[True, True, True]))
        # temp_df['ENSUBJNAME'] = temp_df['ENSUBJNAME'].map(subject_column_replace)
        filter_col = subject_rank_col_common + self.form_rank_col
        temp_df[filter_col].to_excel(rank_writer, sheet_name='subj_rank_form', index=False)

        # 5. Filter Class 1-6 of each subject
        print('sheet', 'subj_rank_group')
        temp_df = pd.merge(self.report_df_dict['score'].view_df[
                               self.report_df_dict['score'].view_df[self.group_rank_col[1]] < 6],
                           self.report_df_dict['student'].view_df[
                               ['REGNO', 'ENNAME', 'CHNAME', 'SEX', 'classlevel', 'class', 'CLASSNO']],
                           how='left', on='REGNO')
        temp_df = (temp_df[~temp_df.subject_key.isin(chi_eng_component_list)]
                   .sort_values(by=['classlevel', 'CROSSCLSSUBGRP', self.group_rank_col[1]],
                                ascending=[True, True, True]))
        filter_col = subject_rank_col_common + ['CROSSCLSSUBGRP'] + self.group_rank_col
        temp_df[filter_col].to_excel(rank_writer, sheet_name='subj_rank_group', index=False)

        rank_writer.save()
        rank_writer.close()


