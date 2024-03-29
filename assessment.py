import os
import pathlib
import openpyxl
import xlutils.copy
import copy
import xlrd
import datetime
import pandas as pd
from jinja2 import Environment
from jinja2 import FileSystemLoader
# need wkhtmltopdf
import pdfkit
from model.examfile import ExamFile
from model.examrun import ExamRun
from PyPDF2 import PdfFileMerger


def converter(column: list):
    return {col: str for col in range(len(column))}


class Assessment:
    def __init__(self, exam_year: str, assessment: str, exam_type: str):

        if assessment in ['ut1', 'ut2', 'exam1', 'exam2', 'mock']:
            pass
        else:
            print(f'assessment must be from [ut1, ut2, exam1, exam2, mock]')
            return

        if exam_type in ['ut1', 'ut2', 'exam1', 'exam2', 'final']:
            pass
        else:
            print(f'exam_type must be from [ut1, ut2, exam1, exam2, final]')
            return

        # assessment_name: year + exam_type
        # ut1 / exam1 / ut2 / exam2 / mock  : { 2021exam/{ assessment }
        # self.assessment_name = assessment
        # self.full_exam = exam_year + 'exam'

        self.exam_year = exam_year
        self.assessment = assessment
        self.exam_type = exam_type

        # assessment_name: year + exam_type
        # 2021ut1 / 2021exam1 / 2021ut2 / 2021exam2 / 2021mock
        self.assessment_name = self.exam_year + assessment

        self.check_mark_template_exam_type = 'exam' if self.assessment in ['exam1', 'exam2', 'mock'] else 'ut'

        # folder setting here
        # python files live in the root level
        # model files live in (root)\model\

        # websams root folder as the ultima root folder
        # home folder : 'websams\2122\'
        # to hold all folders (in the same school year) within home folder

        self.websams_root_folder = os.path.abspath(os.getcwd())                             # /websams
        self.template_root_folder = os.path.join(self.websams_root_folder, 'template')      # /websams/template
        self.schyear_home_folder = os.path.join(self.websams_root_folder, self.exam_year)   # /websams/2021
        # template folder of current year
        # copy from self.template_root_folder
        self.template_folder = os.path.join(self.schyear_home_folder, '_template')          # /websams/2021/_template
        self.html_template_root_folder = os.path.join(self.template_folder, 'html')
        if pathlib.Path(self.template_folder).exists():
            pass
        else:
            print(f'copy { self.template_root_folder } to { self.template_folder }')

        # /websams/2021/2021exam
        self.assessment_root_folder = os.path.join(self.schyear_home_folder, self.exam_year + 'exam')
        if pathlib.Path(self.assessment_root_folder).exists():
            pass
        else:
            os.makedirs(self.assessment_root_folder, exist_ok=True)

        # websams/2021/2021exam/{ exam }
        self.assessment_folder = os.path.join(self.assessment_root_folder, self.assessment)
        # analysis folder to hold all final version analysis files
        self.analysis_folder = os.path.join(self.assessment_root_folder, 'analysis')
        if pathlib.Path(self.analysis_folder).exists():
            pass
        else:
            os.makedirs(self.analysis_folder, exist_ok=True)

        # all sub folders in assessment_folder
        self.merge_folder = os.path.join(self.assessment_folder, 'merge')
        self.check_mark_folder = os.path.join(self.assessment_folder, 'check_mark')
        self.db_folder = os.path.join(self.assessment_folder, 'db')
        self.pdf_folder = os.path.join(self.assessment_folder, 'pdf')
        self.websams_import_folder = os.path.join(self.assessment_folder, 'websams_import')
        self.websams_src_folder = os.path.join(self.assessment_folder, 'websams_src')
        # folder structure inside assessment home
        self.folder = {'websams_src': 'websams_src',
                       'websams_import': 'websams_import',
                       'pdf': 'pdf',
                       'merge': 'merge',
                       'db': 'db',
                       'analysis': 'analysis',
                       'check_mark': 'check_mark'
                       }
        # build folder in assessment home
        # setup assessment mark list
        # for handling of 'final'
        for key, folder in self.folder.items():
            folder_to_create = os.path.join(self.assessment_folder, folder)
            if pathlib.Path(folder_to_create).exists():
                pass
            else:
                print('mkdir', folder_to_create)
                os.makedirs(folder_to_create, exist_ok=True)

        # load exam_run and assessment
        self.exam_run = ExamRun(self.exam_year)
        # load from exam_run
        self.exam_run.load_assessment(self.assessment)
        self.assessment_exam_class_dict = self.exam_run.assessment_dict
        self.ct_dict = self.exam_run.load_ct_to_dict()

        self.assessment_exam_file_dict = {}
        self.assessment_exam_file_tch_dict = {}
        self.assessment_exam_file_state_list = []
        self.assessment_file_path = {}
        self.students = {}
        self.analysis_file_dict = {}
        self.exam_file_state_df = None
        self.summary_df = None

        # file template
        self.view_template_src = os.path.join(self.template_folder, 'exam', 'view-all.xlsm')
        self.analysis_template_src = \
            {'s123': os.path.join(self.template_folder, 'exam', 's123-analysis.xlsm'),
             's456': os.path.join(self.template_folder, 'exam', 's456-analysis.xlsm')}
        # about analysis src file
        # should use exam_run to create the master template for the current school year

        self.check_file = os.path.join(self.db_folder, self.exam_year + '-' + self.assessment + '-view-all.xlsm')
        self.db_file = os.path.join(self.db_folder, self.exam_year + '-' + self.assessment + '-db.xlsx')
        self.view_file = os.path.join(self.db_folder, self.exam_year + '-' + self.assessment + '-view-all.xlsm')

        # 1920examA/db/1920-exam_type-db.xlsx
        # for filtering of mark dataframe into analysis file
        stuinfo = ['key', 'regno', 'enname', 'chname', 'sex', 'classlevel', 'classcode', 'classno']
        subjS123 = ['chi', 'eng', 'mth', 'chs', 'hst', 'geo', 'eco', 'isc', 'phy', 'chm', 'bio']
        subjS123exam1 = ['bik', 'lst']
        subjS123exam2 = ['bik', 'lst', 'pth', 'cps', 'via', 'mus', 'ped', 'dte', 'hec']
        subjS456 = ['chi', 'eng', 'mth', 'm1', 'm2', 'lst', 'chs', 'hst', 'geo', 'eco',
                    'phy', 'chm', 'bio', 'baf', 'ict']
        subjS456exam2 = ['bik', 'mus', 'ped']

        self.analysis_col = {'s123': stuinfo + subjS123 + subjS123exam2,
                             's45': stuinfo + subjS456 + subjS456exam2,
                             's456': stuinfo + subjS456 + subjS456exam2}

        self.analysis_file_dict \
            = {'s123': os.path.join(self.db_folder, self.exam_year + '-' + self.assessment + '-s123-analysis.xlsm'),
               's456': os.path.join(self.db_folder, self.exam_year + '-' + self.assessment + '-s456-analysis.xlsm'),
               }

        # select mark_column by exam
        # for analysis / websams
        if self.assessment in ['ut1', 'ut2']:
            self.mark_col = self.assessment     # ut1, ut2
        elif self.assessment == 'exam1':
            self.mark_col = 'total1'
        elif self.assessment == 'exam2':
            self.mark_col = 'total2'
        elif self.assessment == 'final':
            self.mark_col = 'final'
        elif self.assessment == 'mock':
            self.mark_col = 'final'

        # create ExamFile for each mark file using exam_class_dict (load from exam_run)
        # exam_file_name format change
        # new format
        # class:    { exam_year }_{ assessment }_{ classcode }_{ subject }_{ teacher }.xlsx
        # group:    { exam_year }_{ assessment }_{ classlevel }_{ groupcode}_{ subject }_{ teacher }.xlsx
        for key, exam_class in self.assessment_exam_class_dict.items():
            exam_file_name = self.assessment_name + exam_class.basename
            temp_file_path = os.path.join(self.merge_folder,
                                          exam_class.subject,
                                          exam_file_name)
            self.assessment_exam_file_dict[key] = ExamFile(exam_class,
                                                           temp_file_path,
                                                           self.assessment_name)

    def __str__(self):
        return self.exam_year + '/' + self.assessment + '/' + self.exam_type

    def create_markfile(self):
        self.exam_run.create()

    def rename_markfile(self):
        # wrapper for ExamRun.rename
        return 0

    def check_assessment_file(self):
        # check exam files in assessment folder
        is_ready = 0
        for key, exam_file in self.assessment_exam_file_dict.items():
            temp = 1 - exam_file.check_file()
            is_ready += temp
        if is_ready == 0:
            print('All files are ready.')

    def exam_file_to_dict(self):
        # logfile = open('log.txt', 'w')
        # get back exam_type from exam_col first
        # exam_type = exam_from_col[exam_col]

        # build 3 items:
        # student <--- for websams file merge
        # temp_dic
        # markrow

        print('********************************************************************')
        print('***                                                              ***')
        print('***                          Warning!                            ***')
        print('***                                                              ***')
        print('********************************************************************')
        print('The excel files shall be freshly saved by excel, not by openpyxl.   ')
        print('Otherwise, no values (by formula) can be read.')
        print('\n')
        print('read mark file and merge into a dict.')
        print('mark file db column taken:'.format(self.mark_col))

        # match markfile db sheet header
        db_header = ['regno', 'enname', 'chname', 'sex',  # 0-3
                     'classlevel', 'classcode', 'classno', 'subject', 'group',  # 4-8
                     'ut1', 'daily1', 'exam1', 'total1',  # 9-12
                     'ut2', 'daily2', 'exam2', 'total2', 'final',  # 13-17
                     't1comp1', 't1comp2', 't1comp3', 't1comp4', 't1comp5',  # 18-22
                     't2comp1', 't2comp2', 't2comp3', 't2comp4', 't2comp5',  # 23-27
                     'fcomp1', 'fcomp2', 'fcomp3', 'fcomp4', 'fcomp5',  # 28-32
                     ]
        # match view file (for check mark) header
        view_header = ['key1', 'key1x', 'key2', 'formkey', 'key3', 'xgroup', 'examtype', 'score',  # 1
                       'regno', 'enname', 'chname', 'sex',  # 9
                       'classlevel', 'classcode', 'classno', 'subject', 'group',  # 13
                       'ut1', 'daily1', 'exam1', 'total1',  # 18
                       'ut2', 'daily2', 'exam2', 'total2', 'final',  # 23
                       't1comp1', 't1comp2', 't1comp3', 't1comp4', 't1comp5',  # 28
                       't2comp1', 't2comp2', 't2comp3', 't2comp4', 't2comp5',  # 33
                       'fcomp1', 'fcomp2', 'fcomp3', 'fcomp4', 'fcomp5',  # 38
                       ]

        comp_col = {'ut1': [0],
                    'total1': ['t1comp1', 't1comp2', 't1comp3', 't1comp4', 't1comp5'],
                    'ut2': [0],
                    'total2': ['t2comp1', 't2comp2', 't2comp3', 't2comp4', 't2comp5'],
                    'final': ['fcomp1', 'fcomp2', 'fcomp3', 'fcomp4', 'fcomp5'],
                    'mock': ['fcomp1', 'fcomp2', 'fcomp3', 'fcomp4', 'fcomp5'],
                    }

        # db_header_dict = dict(zip(db_header, range(1, len(db_header)+1)))
        # use a new wb to save db
        exam_db_wb = openpyxl.Workbook(write_only=True)
        exam_db_ws = exam_db_wb.create_sheet(title='db')

        # insert header
        # append markrow afterwards
        exam_db_ws.append(db_header)

        # loop all folders in merge/(exam_type)/
        # use students dict to build student_dict dataframe
        # also use for websams import files
        student_mark = []       # for exam-view file: temp_dic
        self.students = {}
        num_of_files = 0

        # define the path: merge_folder
        # loop through all the subject folder in merge/(exam) folder
        # original_file_list = [self.full_exam + item for item in self.assessment_exam_file_dict.keys()]
        # current_directory = pathlib.Path(self.merge_folder)
        exam_file_state_column = ['classlevel', 'classcode', 'class_type', 'groupcode', 'subject',
                                  'subj_code', 'subj_key', 'path', 'tch', 'teacher',
                                  'room', 'exam_file_name', 'exam_file_path', 'exam_file_state']
        statistics_item_list1 = ['No of Ss', 'No of Pass', 'NoFail', 'NoZero', 'Passing%',
                                 'Mean', 'SD', 'Max', 'Q3', 'Q2', 'Q1', 'Min']
        dump_stat_list = [-1]*len(statistics_item_list1)
        for key, exam_file in self.assessment_exam_file_dict.items():
            # print('{}: {}'.format(key, exam_file.subj_key))
            if key[:2] in ['s1', 's2', 's3', 's4', 's5', 's6']:
                # check file and update exam_file.file_state
                exam_file.check_file()
                if exam_file.file_state:
                    exam_file_dict = exam_file.to_dict()
                    class_stat_dict = exam_file.get_class_statistics()

                    # append exam_file state to a list
                    self.assessment_exam_file_state_list.append(exam_file.file_state_list()
                                                                + list(class_stat_dict.values()))

                    print('\t{}: {}'.format(key, exam_file.file_path))
                    for student in exam_file_dict:
                        if student['regno'] is not None:
                            regno = student['regno']
                            if regno not in self.students:
                                self.students[regno] = {'key': student['classcode']
                                                        + str(int(student['classno'])).zfill(2),
                                                        'regno': student['regno'],
                                                        'enname': student['enname'],
                                                        'chname': student['chname'],
                                                        'sex': student['sex'],
                                                        'classlevel': student['classlevel'],
                                                        'classcode': student['classcode'],
                                                        'classno': student['classno'],
                                                        }

                            subject = student['subject'].strip()
                            # subject if x1phy
                            # use exam_type to select the exam column
                            # in new setting subject and xgroup code are separated
                            if len(subject) <= 3:      # non-x subject
                                self.students[regno][subject] = student[self.mark_col]
                                # handle chi eng components
                                if subject in ['chi', 'eng']:
                                    if self.assessment in ['exam1', 'exam2', 'final', 'mock']:
                                        # need to handle components
                                        self.students[regno][subject+'01'] = student[comp_col[self.mark_col][0]]
                                        self.students[regno][subject+'02'] = student[comp_col[self.mark_col][1]]
                                        self.students[regno][subject+'03'] = student[comp_col[self.mark_col][2]]
                                        self.students[regno][subject+'04'] = student[comp_col[self.mark_col][3]]
                                        self.students[regno][subject+'05'] = student[comp_col[self.mark_col][4]]
                                    elif self.assessment in ['ut1', 'ut2']:
                                        # only 1 component in ut
                                        self.students[regno][subject + '01'] = student[self.mark_col]
                                    else:
                                        pass
                                else:
                                    pass
                            else:
                                # x-subject: trim x1
                                # should be no need any more
                                self.students[regno][subject[2:]] = student[self.mark_col]  # ut2

                            student_mark.append(student)
                            exam_db_ws.append(list(student.values()))
                            # print(list(student.values()))
                            # print(list(student.keys()))
                            # print(student)
                    num_of_files += 1

                else:
                    print('{}: {} not ok.'.format(exam_file.teacher, exam_file.file_path))
                    self.assessment_exam_file_state_list.append(exam_file.file_state_list() + dump_stat_list)

        self.exam_file_state_df = pd.DataFrame(self.assessment_exam_file_state_list,
                                               columns=exam_file_state_column + statistics_item_list1)

        exam_file_save = 'exam_file_state.xlsx'
        self.exam_file_state_df.to_excel(exam_file_save, index=False)
        print('\texam_file_state file: {}'.format(exam_file_save))

        # print(student_mark)
        # input()
        # save exam db file
        print('\tmerged Files: {}'.format(num_of_files))
        print('\tnumber of Score records: {}'.format(len(student_mark)))
        print('\tdb file: {}'.format(self.db_file))
        exam_db_wb.save(self.db_file)

        print('\tnow, generate analysis file.')
        print('\tshall use the vba of generated file to print all analysis report.')

        # then use excel macro to dump class analysis in pdf
        # code to dump values to 1 master analysis file
        # write to exam-check file
        # then use crystal report to generate all check marks files
        view_wb = openpyxl.load_workbook(filename=self.view_template_src, keep_vba=True)
        score_ws = view_wb['score']

        # dump Header
        # loop df1 (junior)
        row = 2
        # print(student_mark[0])
        for record in student_mark:
            # print(record)
            # k = len(record)
            for key, item in record.items():
                # a = input()
                if type(key) is str:
                    if key.lower() in view_header:
                        score_ws.cell(row=row, column=view_header.index(key.lower()) + 1).value = item
                        # if key.lower() == 'group':
                        # print(key, score_ws.cell(row=row, column=view_header.index(key.lower()) + 1).value)
                    else:
                        pass
                        # print('subject:{} - {} : key not found!'.format(record['subject'], key))

                else:
                    print('{}: {} - Non-str'.format(key, item))

            row += 1
        view_wb.save(self.view_file)
        view_wb.close()
        print('check mark file: ' + self.view_file)

        # create df of all student record of current exam
        # for filtering into S123 and S456
        # for writing into analysis file
        # !! work around
        # produce a dummy row with all subjects
        # such that the dataframe will be full header for filtering
        self.students['CDG000001'] = {'key': 'S3Z50',
                                      'regno': 'CDG000001',
                                      'enname': 'xxxxxx',
                                      'chname': 'xxxxxx',
                                      'sex': 'x',
                                      'classlevel': 'S3',
                                      'classcode': 'S3Z',
                                      'classno': 50,
                                      'chi': 0,
                                      'eng': 0,
                                      'mth': 0,
                                      'm1': 0,
                                      'm2': 0,
                                      'lst': 0,
                                      'chs': 0,
                                      'hst': 0,
                                      'geo': 0,
                                      'eco': 0,
                                      'baf': 0,
                                      'isc': 0,
                                      'phy': 0,
                                      'chm': 0,
                                      'bio': 0,
                                      'ict': 0,
                                      'bik': 0,
                                      'cps': 0,
                                      'pth': 0,
                                      'via': 0,
                                      'mus': 0,
                                      'ped': 0,
                                      'dte': 0,
                                      'hec': 0,
                                      }

        self.students['CDG000002'] = {'key': 'S6Z50',
                                      'regno': 'CDG000001',
                                      'enname': 'xxxxxx',
                                      'chname': 'xxxxxx',
                                      'sex': 'x',
                                      'classlevel': 'S6',
                                      'classcode': 'S6Z',
                                      'classno': 50,
                                      'chi': 0,
                                      'eng': 0,
                                      'mth': 0,
                                      'm1': 0,
                                      'm2': 0,
                                      'lst': 0,
                                      'chs': 0,
                                      'hst': 0,
                                      'geo': 0,
                                      'eco': 0,
                                      'baf': 0,
                                      'isc': 0,
                                      'phy': 0,
                                      'chm': 0,
                                      'bio': 0,
                                      'ict': 0,
                                      'bik': 0,
                                      'cps': 0,
                                      'pth': 0,
                                      'via': 0,
                                      'mus': 0,
                                      'ped': 0,
                                      'dte': 0,
                                      'hec': 0,
                                      }
        # student_dict
        # return exam_col in future
        # print(self.students)
        return self.students

    # for merge websams files from the given student mark dict
    def merge_websams(self):
        # use it to merge to websams file by
        # loop through all websams source file
        # set the path to websams xls import folder
        websams_subject_code = {'045_BIO_E_Score_01233210045': 'bio',
                                '070_CHEM_E_Score_01233210070': 'chm',
                                '075_CHIS_C_Score_01233210075': 'chs',
                                '080_CHIN_C_Score_01233210080': 'chi',
                                '080_CHIN_Read_C_Score_0123321008001': 'chi01',
                                '080_CHIN_Writ_C_Score_0123321008002': 'chi02',
                                '080_CHIN_List_C_Score_0123321008003': 'chi03',
                                '080_CHIN_Spea_C_Score_0123321008004': 'chi04',
                                '080_CHIN_Dict_C_Score_0123321008005': 'chi05',
                                '110_CL_E_Score_01233210110': 'cps',
                                '12N_Acc_E_Score_0123321012N': 'baf',
                                '135_ECON_E_Score_01233210135': 'eco',
                                '165_ENG_E_Score_01233210165': 'eng',
                                '165_ENG_Read_E_Score_0123321016501': 'eng01',
                                '165_ENG_Writ_E_Score_0123321016502': 'eng02',
                                '165_ENG_List_E_Score_0123321016503': 'eng03',
                                '165_ENG_Spea_E_Score_0123321016504': 'eng04',
                                '185_E&RE_C_Score_01233210185': 'bik',
                                '210_GEOG_E_Score_01233210210': 'geo',
                                '280_MATH_E_Score_01233210280': 'mth',
                                '22S_MACO_E_Score_0123321022S': 'mth',
                                '235_HIST_E_Score_01233210235': 'hst',
                                '23S_MAM1_E_Score_0123321023S': 'm1',
                                '24S_MAM2_E_Score_0123321024S': 'm2',
                                '265_LIBS_C_Score_01233210265': 'lst',
                                '300_MUS_E_Score_01233210300': 'mus',
                                '310_PE_E_Score_01233210310': 'ped',
                                '315_PHY_E_Score_01233210315': 'phy',
                                '81N_ICT_E_Score_0123321081N': 'ict',
                                '260_SCJ_E_Score_01233210260': 'isc',
                                '130_D&T_E_Score_01233210130': 'dte',
                                '240_HEC_E_Score_01233210240': 'hec',
                                '350_PTH_O_Score_01233210350': 'pth',
                                '350_PTH_C_Score_01233210350': 'pth',
                                '432_VA_E_Score_01233210432': 'via',
                                }

        current_directory = pathlib.Path(self.websams_src_folder)
        print(current_directory)
        print('start generating websams files to', self.websams_import_folder)
        # read xls using xlrd
        # loop all files in websams import folder
        for current_file in current_directory.iterdir():
            # DE_19056020190329_013_3_3_S4_S4B.xls
            # print('current file:', current_file)
            # should check exam_type and exam_col in future
            if current_file.name[0:2] == 'DE':
                current_file_item = {'prefix': 'DE_',
                                     'school_code': current_file.name[3:9],
                                     'dateVal': current_file.name[9:17],
                                     'idx': current_file.name[18:21],
                                     'classlevel': current_file.name[26:28],
                                     'classcode': current_file.name[29:32]
                                     }
                print(current_file)
                websams_file_src = xlrd.open_workbook(current_file)
                copy_sheet = websams_file_src.sheet_by_index(0)
                # copy xls file to enable write
                temp_wb = xlutils.copy.copy(websams_file_src)
                # loop through all student in xls import file
                for row in range(1, len(copy_sheet.col_values(0))):
                    # print(copySheet.cell(row, 8).value[1:])
                    cur_regno = copy_sheet.cell(row, 8).value[1:]
                    # check student in all_stu_dict
                    if cur_regno in self.students:
                        cur_student = self.students[cur_regno]
                        # print(curStudent)
                        # loop through all subjects in xls import file
                        for col in range(len(copy_sheet.row_values(0))):
                            cur_subject = copy_sheet.row_values(0)[col][5:]
                            # print(curSubject)
                            # check subject in student marks dict
                            if cur_subject in websams_subject_code:
                                subject_code = websams_subject_code[copy_sheet.row_values(0)[col][5:]]
                                # print(subject_code)
                                if subject_code in cur_student:
                                    # print(cur_student[subject_code])
                                    temp_wb.get_sheet(0).write(row, col, cur_student[subject_code])
                                    # print(cur_regno, subject_code, cur_student[subject_code])
                    else:
                        # wrong reference of name as cur_student does not exist
                        print('{}:{} does not exist.'.format(current_file_item['classcode'], cur_regno))
                save_file = os.path.join(self.websams_import_folder, current_file.name)
                temp_wb.save(save_file)
                # print('\t' + current_file_item['classcode'] + ': ' + save_file)

        print('all websams files:', self.websams_import_folder)

    def students_df(self):
        temp_dict = []  # for later to convert to df
        for regno in self.students:
            temp_dict.append(self.students[regno])
        return pd.DataFrame(temp_dict)

    def write_analysis(self):
        # filter junior form and sort by classlevel-classcode-classno
        # skip junior for S6 mock
        # check exam_type
        # if exam2 then also gen final analysis
        # for filtering of mark dataframe into analysis file
        if self.assessment in ['ut1', 'exam1', 'ut2', 'exam2', 'final']:
            classlevel_list = ['s123', 's456']
        else:
            classlevel_list = ['s456']

        class_filter_list = {'s123': ['S1', 'S2', 'S3'],
                             's456': ['S4', 'S5', 'S6'],
                             's45': ['S4', 'S5'],
                             's6': ['S6']}

        print('write analysis')
        df = self.students_df()
        # 'classlevel': ['s123', 's456']
        for classlevel in classlevel_list:

            classlevel_filter = class_filter_list[classlevel]
            column_full = self.analysis_col[classlevel]

            # try:
            temp_df = df[df.classlevel.isin(classlevel_filter)].sort_values(
                by=['classlevel', 'classcode', 'classno'],
                ascending=[True, True, True])
            analysis_wb = openpyxl.load_workbook(filename=self.analysis_template_src[classlevel], keep_vba=True)
            src_ws = analysis_wb['src']
            # dump Full Header
            for col_title in column_full:
                src_ws.cell(row=1, column=column_full.index(col_title) + 1).value = col_title

            row = 2
            for student in temp_df.to_dict('records'):
                # print(student)
                for key, item in student.items():
                    if key in column_full:
                        src_ws.cell(row=row, column=column_full.index(key) + 1).value = item
                    else:
                        pass
                    #    print('Key({}) not found!'.format(key))
                row += 1

            analysis_wb.save(self.analysis_file_dict[classlevel])
            analysis_wb.close()
            # self.analysis_file_list[analysis][classlevel] = analysis_file
            print('{} analysis: {}'.format(classlevel, self.analysis_file_dict[classlevel]))

            # except:
            #    print('df: some files are missing.')

    def copy_analysis(self):

        analysis_file_dict = \
            {'s123': os.path.join(self.analysis_folder,
                                  self.exam_year + '-' + self.assessment + '-s123-analysis_Final.xlsm'),
             's456': os.path.join(self.analysis_folder,
                                  self.exam_year + '-' + self.assessment + '-s456-analysis_Final.xlsm'),
             }

        copy_ws_list = {'ut1': ['s1ut', 's2ut', 's3ut', 's4ut', 's5ut', 's6ut'],
                        'exam1': ['s1ex1', 's2ex1', 's3ex1', 's4ex1', 's5ex1'],
                        'ut2': ['s1ut', 's2ut', 's3ut', 's4ut', 's5ut'],
                        'exam2': ['s1ex2', 's2ex2', 's3ex2', 's4ex2', 's5ex2'],
                        'final': ['s1final', 's2final', 's3final', 's4final', 's5final'],
                        }

        if pathlib.Path(analysis_file_dict['s123']).exists() and pathlib.Path(analysis_file_dict['s456']).exists:
            # create a new workbook
            wb_new = openpyxl.Workbook()

            for classlevel, file_open in analysis_file_dict.items():
                print('use analysis file:', file_open)
                analysis_wb_src = openpyxl.load_workbook(filename=file_open, data_only=True)
                sheet_list = analysis_wb_src.sheetnames
                sheet_list.sort()
                print(sheet_list)
                for sheet_name in sheet_list:
                    print('sheet_name:', sheet_name)
                    if sheet_name in copy_ws_list[self.assessment]:
                        print('\tsheet', sheet_name, 'copying')
                        ws_src = analysis_wb_src[sheet_name]
                        # ws_new = wb_new.active
                        ws_new = wb_new.create_sheet(title=sheet_name)

                        # copy worksheet attributes
                        ws_new.sheet_format = copy.copy(ws_src.sheet_format)
                        ws_new.sheet_format = copy.copy(ws_src.sheet_format)
                        ws_new.sheet_properties = copy.copy(ws_src.sheet_properties)
                        ws_new.merged_cells = copy.copy(ws_src.merged_cells)
                        ws_new.page_margins = copy.copy(ws_src.page_margins)
                        ws_new.page_setup = copy.copy(ws_src.page_setup)
                        ws_new.print_options = copy.copy(ws_src.print_options)
                        ws_new.row_dimensions = copy.copy(ws_src.row_dimensions)
                        ws_new.column_dimensions = copy.copy(ws_src.column_dimensions)
                        ws_new._print_area = copy.copy(ws_src._print_area)

                        # copy cell by cell
                        for row in ws_src.rows:
                            for cell in row:
                                new_cell = ws_new.cell(row=cell.row, column=cell.col_idx, value=cell.value)
                                if cell.has_style:
                                    new_cell.font = copy.copy(cell.font)
                                    new_cell.border = copy.copy(cell.border)
                                    new_cell.fill = copy.copy(cell.fill)
                                    new_cell.number_format = copy.copy(cell.number_format)
                                    new_cell.protection = copy.copy(cell.protection)
                                    new_cell.alignment = copy.copy(cell.alignment)

            wb_file_save = os.path.join(self.analysis_folder,
                                        self.exam_year + '-' + self.assessment + '-analysis-value.xlsx')
            wb_new.save(wb_file_save)
            wb_new.close()
            print('{} created'.format(wb_file_save))
        else:
            print('{} or {} is missing'.format(analysis_file_dict['s123'], analysis_file_dict['s456']))

    def db_to_pdf(self, exam_file: ExamFile, file_rank: str):

        # need to choose exam template by exam
        exam_file_state_column = ['classlevel', 'classcode', 'class_type', 'groupcode', 'subject',
                                  'subj_code', 'subj_key', 'path', 'tch', 'teacher',
                                  'room', 'exam_file_name', 'exam_file_path', 'exam_file_state']
        pass_mark = exam_file.pass_mark
        cat = exam_file.cat

        check_mark_template = 'check_mark_' + self.check_mark_template_exam_type + '_' + cat + '_template.html'
        print(check_mark_template)
        # exam_file load mark_col(s) by its filename
        class_statistics = exam_file.get_class_statistics()
        student_marks_dict = exam_file.db_to_print().to_dict('records')
        # print(student_marks_dict)
        # build a pass column
        # for student in student_marks_dict:
        #     student['pass'] = (float(student['mark']) >= pass_mark)

        stat_dict = {'No. of Students:': class_statistics['No of Ss'],
                     'No. of Passed (%):': '{} ({}%)'.format(class_statistics['No of Pass'],
                                                             class_statistics['Passing%']),
                     'No. of Failed:': class_statistics['NoFail'],
                     'No. of 0 Mark:': class_statistics['NoZero'],
                     'Mean:': class_statistics['Mean'],
                     'Maximum:': class_statistics['Max'],
                     'Upper Quartile:': class_statistics['Q3'],
                     'Median:': class_statistics['Q2'],
                     'Lower Quartile:': class_statistics['Q1'],
                     'Minimum:': class_statistics['Min'],
                     }

        # exam_file_dict = dict(zip(exam_file_state_column, exam_file.file_state_list()))
        exam_file_dict = {'school_year': '20{}/20{}'.format(self.exam_year[:2], self.exam_year[2:]),
                          'classlevel': exam_file.classlevel.upper(),
                          'classcode': exam_file.classcode.upper(),
                          'subject': exam_file.subject.title(),
                          'teacher': exam_file.teacher.upper(),
                          'exam_type': self.assessment.title(),
                          'file_code': '{}-{}'.format(exam_file.classcode.upper(),
                                                      exam_file.subject.upper()),
                          'tch_file_rank': '{}({})'.format(exam_file.teacher.upper(), file_rank),
                          'file_rank': file_rank,
                          }

        filename = exam_file.exam_file_name[:-5]
        print_string = 'Print: {} ({})'.format(datetime.datetime.now().strftime('%Y-%m-%d'),
                                               datetime.datetime.now().strftime('%H:%M'))

        # template: (root)/template/html/check_mark/
        html_check_mark_folder = os.path.join(self.html_template_root_folder, 'check_mark')
        # need to add code to select template by assessment
        # chi / eng / normal / (daily only)

        css = os.path.join(html_check_mark_folder, 'check_mark.css')
        print(html_check_mark_folder)
        env = Environment(loader=FileSystemLoader(html_check_mark_folder))
        env.globals['enumerate'] = enumerate
        template = env.get_template(check_mark_template)

        template_vars = {'print_date': print_string,
                         'exam_file': exam_file_dict,
                         'students': student_marks_dict,
                         'statistics': stat_dict,
                         }

        html_out = template.render(template_vars)
        html_save = os.path.join(self.check_mark_folder, 'temp.html')
        pdf_save = os.path.join(self.check_mark_folder, filename + '_check.pdf')

        options = {
            'page-size': 'A4',
            'margin-top': '1.0cm',
            'margin-right': '1.0cm',
            'margin-bottom': '0.5cm',
            'margin-left': '1.0cm',
            'encoding': "UTF-8",
            'no-outline': None,
            'enable-local-file-access': None
            }

        with open(html_save, 'w', encoding='utf-8') as f:
            f.write(html_out)
            f.close()
        pdfkit.from_file(html_save, pdf_save, css=css, options=options)

        # print('{}: {} saved.'.format('db', html_save))
        print('{}: {} saved.'.format('db', pdf_save))
        return pdf_save

    def get_exam_file_summary(self):
        temp_df = self.exam_file_state_df[['teacher', 'exam_file_state', 'room']]
        temp_df.loc[:, 'tch'] = temp_df['teacher']
        summary_df = temp_df.groupby(['teacher', 'room'], as_index=False).agg({'exam_file_state': 'sum',
                                                                               'tch': 'count'})
        summary_df['file'] = summary_df['exam_file_state'].astype(str) + '/' + summary_df['tch'].astype(str)
        self.summary_df = summary_df[['teacher', 'room', 'tch', 'file']].sort_values(by=['room', 'teacher'],
                                                                                     ascending=[True, True])

    def check_mark(self):
        if self.exam_file_state_df is None:
            print('Exam files are not yet loaded.')
        else:
            # self.summary_df.to_excel('summary.xlsx')
            # check_mark_file_list = []

            pdf_merger = PdfFileMerger()
            file_handles = []
            env = Environment(loader=FileSystemLoader(self.html_template_root_folder))
            options = {
                'page-size': 'A4',
                'margin-top': '1.0cm',
                'margin-right': '1.0cm',
                'margin-bottom': '0.5cm',
                'margin-left': '1.0cm',
                'encoding': "UTF-8",
                'no-outline': None,
                'enable-local-file-access': None
            }
            # get summary
            self.get_exam_file_summary()

            html_save = os.path.join(self.check_mark_folder, 'temp.html')
            summary_save = os.path.join(self.check_mark_folder, 'summary_check.pdf')
            with open(html_save, 'w', encoding='utf-8') as f:
                f.write(self.summary_df.to_html())
                f.close()
            pdfkit.from_file(html_save, summary_save, options=options)
            pdf_merger.append(summary_save)

            # subj_key ({class_code}{subject_code}) as the unique key
            self.exam_file_state_df['key'] = self.exam_file_state_df['subj_key']
            self.exam_file_state_df.set_index('key', inplace=True)
            self.exam_file_state_df.sort_values(by=['room', 'tch', 'subj_key'],
                                                ascending=[True, True, True], inplace=True)

            tch_list = list(set(self.exam_file_state_df['tch'].to_list()))
            tch_list.sort()
            file_rank_dict = {}
            for tch in tch_list:
                temp_dict = (self.exam_file_state_df[self.exam_file_state_df.tch == tch]['subj_key']
                             .rank()
                             .astype(int)
                             .to_dict())
                m = max(temp_dict.values())
                file_rank_dict.update({k: str(n) + "/" + str(m) for k, n in temp_dict.items()})
            # print(file_rank_dict)
            # self.exam_file_state_df.to_excel('exam_file_state.xlsx')
            exam_file_state_dict = self.exam_file_state_df.to_dict('index')
            # print(exam_file_state_dict)

            for n, (key, exam_file_state) in enumerate(exam_file_state_dict.items(), start=1):
                # restrict n to debug
                if n <= 500:  # 1000
                    if self.assessment_exam_file_dict[key].check_file():
                        temp_path = self.db_to_pdf(self.assessment_exam_file_dict[key], file_rank_dict[key])
                        print('#{}-{}: {}'.format(n, key, temp_path))
                        # check_mark_file_list.append(temp_path)
                        pdf_merger.append(temp_path)

            merged_check_mark_file = os.path.join(self.assessment_folder, 'check_mark_file.pdf')
            with open(merged_check_mark_file, 'wb') as file_object:
                pdf_merger.write(file_object)
            print('check files generated: {}'.format(merged_check_mark_file))

    def generate_junior_analysis(self):

        junior_class_dict = {'s1': ['S1A', 'S1B', 'S1C', 'S1D'],
                             's2': ['S2A', 'S2B', 'S2C', 'S2D'],
                             's3': ['S3A', 'S3B', 'S3C', 'S3D']}

        mark_col_ut_dict = {'s1': ['chi', 'eng', 'mth', 'chs', 'geo', 'isc'],
                            's2': ['chi', 'eng', 'mth', 'chs', 'hst', 'geo', 'isc'],
                            's3': ['chi', 'eng', 'mth', 'chs', 'hst', 'geo', 'eco', 'phy', 'chm', 'bio']}

        stat_col_type_list = ['No of Ss', 'Passing%', '0 - 10 (excluding 10)', '10 - 20 (excluding 20)',
                              '20 - 30 (excluding 30)', '30 - 40 (excluding 40)', '40 - 50 (excluding 50)',
                              '50 - 60 (excluding 60)', '60 - 70 (excluding 70)', '70 - 80 (excluding 80)',
                              '80 - 90 (excluding 90)', '90 - 100']

        print_wm_header_dict = {'WM': 'wm', 'Class Rank': 'rank', 'Form Rank': 'rank', '及格科目': 'rank'}

        pdf_merger_master = PdfFileMerger()
        pdf_merger_ct = PdfFileMerger()

        junior_analysis_file = '2021-final-s123-analysis_Final.xlsm'
        junior_analysis_file = self.analysis_file_dict['s123']

        exam_type = junior_analysis_file.split('-')[1]
        junior_analysis_open = os.path.join(self.db_folder, junior_analysis_file)
        print(junior_analysis_open)

        student_df = pd.read_excel(junior_analysis_open, sheet_name='print')
        statistics_df = pd.read_excel(junior_analysis_open, sheet_name='stat')

        passmark, top = 50, 30
        for classlevel, classlist in junior_class_dict.items():
            for classcode in classlist:

                class_df = student_df[student_df['classcode'] == classcode].copy()
                statistics_col = ['item'] + mark_col_ut_dict[classlevel] + ['wm']
                stat_temp_df = statistics_df[statistics_df['classcode'] == classcode][statistics_col]
                class_stat_dict = stat_temp_df.to_dict('records')

                # stu_g   chi_g	eng_g	mth_g	chs_g	hst_g   geo_g	eco_g	phy_g	chm_g	bio_g   wm_g
                print_stuinfo = ['stuinfo_g']
                class_df['stuinfo_g'] = [[x1, x2, x3] for x1, x2, x3 in zip(class_df['key'],
                                                                            class_df['chname'],
                                                                            class_df['sex'])]

                # need class student counter
                print_stu_header = ['Class/No.', 'Chi Name', 'sex']
                print_stu_header_dict = {'Class/No.': 'key', 'Chi Name': 'chname', 'Sex': 'sex'}
                print_subject_header_dict = {}
                print_subject = [col + '_g' for col in mark_col_ut_dict[classlevel]]

                # build subject item group
                for subject in mark_col_ut_dict[classlevel]:
                    rank_col = subject.title() + ' Form Rank'
                    # rank_col = 'Form Rank'
                    # add a selector for bik grade later
                    if subject in mark_col_ut_dict[classlevel]:
                        # print(subject)
                        print_subject_header_dict[subject.title()] = '"mark"'
                        print_subject_header_dict[rank_col] = '"rank"'
                        # build subject item group
                        class_df[subject + '_s'] = class_df[subject].apply(lambda x: '{:.2f}'.format(x))
                        class_df[subject + '_p'] = class_df[subject].apply(
                            lambda x: '"mark"' if x >= passmark else '"mark fail"')
                        class_df[rank_col + '_s'] = class_df[rank_col].astype(int)
                        class_df[rank_col + '_p'] = class_df[rank_col].apply(
                            lambda x: '"rank top"' if x <= top else '"rank"')
                        # build subject group
                        class_df[subject + '_g'] = [[x1, x2, x3, x4] for x1, x2, x3, x4 in
                                                    zip(class_df[subject+'_s'], class_df[subject+'_p'],
                                                        class_df[rank_col+'_s'], class_df[rank_col+'_p'])]
                    else:
                        pass

                # build wm item group
                class_df['wm_p'] = class_df['wm'].apply(lambda x: '"mark"' if x >= passmark else '"mark fail"')
                class_df['crank_p'] = class_df['Class Rank'].apply(lambda x: '"rank top"' if x <= 10 else '"rank"')
                class_df['frank_p'] = class_df['Form Rank'].apply(lambda x: '"rank top"' if x <= top else '"rank"')
                class_df['wm_g'] = [[x1, x2, x3, x4, x5, x6, x7] for x1, x2, x3, x4, x5, x6, x7 in zip(
                    class_df['wm'].apply(lambda x: '{:.2f}'.format(x)),
                    class_df['wm_p'],
                    class_df['Class Rank'].astype(int),
                    class_df['crank_p'],
                    class_df['Form Rank'].astype(int),
                    class_df['frank_p'],
                    class_df['及格科目'].astype(int))]

                print_wm = ['wm_g']
                print_col = print_stuinfo + print_subject + print_wm
                # class_df.to_excel('test.xlsx', index=False)
                # class_df[print_col].to_excel('print.xlsx', index=False)

                student_dict = class_df[print_col].to_dict('records')

                # setup for jinja2
                html_analysis_folder = os.path.join(self.html_template_root_folder, 'analysis')
                env = Environment(loader=FileSystemLoader(html_analysis_folder))
                template = env.get_template('analysis_template_junior_ut.html')
                css = os.path.join(html_analysis_folder, 'analysis_template_junior.css')

                print_date = 'Print: {} ({})'.format(datetime.datetime.now().strftime('%Y-%m-%d'),
                                                     datetime.datetime.now().strftime('%H:%M'))
                class_info_dict = self.ct_dict[classcode.lower()]
                class_info_dict['exam_year'] = self.exam_year
                class_info_dict['exam_name'] = exam_type.title()
                class_info_dict['classcode'] = classcode.upper()
                class_info_dict['print_date'] = print_date

                template_vars = {'students': student_dict,
                                 'class_dict': class_info_dict,
                                 'statistics': class_stat_dict,
                                 'statistics_header': statistics_col,
                                 'statistics_col': stat_col_type_list,
                                 'print_stu_header': print_stu_header_dict,
                                 'print_subject_header': print_subject_header_dict,
                                 'print_wm_header': print_wm_header_dict,
                                 'print_subjects': print_subject,
                                 'print_stuinfo': print_stuinfo,
                                 'print_wm': print_wm,
                                 'css': css,
                                 }

                # print(print_stu_header_dict)
                # print(print_subject_header_dict)
                # print(print_wm_header_dict)

                options = {'page-size': 'A4',
                           'orientation': 'landscape',
                           'margin-top': '1.0cm',
                           'margin-right': '1.0cm',
                           'margin-bottom': '0.5cm',
                           'margin-left': '1.0cm',
                           'encoding': "UTF-8",
                           # 'encoding': 'ISO-8859-1',
                           'no-outline': None,
                           'enable-local-file-access': None}

                html_out = template.render(template_vars)
                html_save = os.path.join(self.db_folder, 'dump', 'temp.html')
                pdf_save = os.path.join(self.db_folder, 'dump',
                                        self.exam_year + '_' + exam_type + '_'
                                        + self.assessment + '_' + classcode.lower() + '.pdf')

                with open(html_save, 'w', encoding='utf-8') as f:
                    f.write(html_out)
                    f.close()

                pdfkit.from_file(html_save, pdf_save, css=css, options=options)
                print('{}: {} saved.'.format(classcode, pdf_save))

                pdf_merger_master.append(pdf_save)
                pdf_merger_ct.append(pdf_save)
                pdf_merger_ct.append(pdf_save)

        analysis_file_master = os.path.join(self.assessment_folder, 'pdf',
                                            self.exam_year + '_' + exam_type + '_' + 'analysis_junior.pdf')
        analysis_file_ct = os.path.join(self.assessment_folder, 'pdf',
                                        self.exam_year + '_' + exam_type + '_' + 'analysis_junior_ct.pdf')

        with open(analysis_file_master, 'wb') as file_object:
            pdf_merger_master.write(file_object)

        with open(analysis_file_ct, 'wb') as file_object:
            pdf_merger_ct.write(file_object)

        print(f'class analysis pdf generated: { analysis_file_master }')

    def generate_senior_analysis(self):

        senior_class_dict = {'s4': ['S4A', 'S4B', 'S4C', 'S4D'],
                             's5': ['S5A', 'S5B', 'S5C', 'S5D'],
                             's6': ['S6A', 'S6B', 'S6C', 'S6D']}

        mark_col_ut_dict = {'s4': ['chi', 'eng', 'mth', 'm1', 'm2', 'lst',
                                   'chs', 'hst', 'geo', 'eco', 'phy', 'chm', 'bio', 'baf', 'ict'],
                            's5': ['chi', 'eng', 'mth', 'm1', 'm2', 'lst',
                                   'chs', 'hst', 'geo', 'eco', 'phy', 'chm', 'bio', 'baf', 'ict'],
                            's6': ['chi', 'eng', 'mth', 'm1', 'm2', 'lst',
                                   'chs', 'hst', 'geo', 'eco', 'phy', 'chm', 'bio', 'baf', 'ict']}

        stat_col_type_list = ['No of Ss', 'Passing%', '0 - 10 (excluding 10)', '10 - 20 (excluding 20)',
                              '20 - 30 (excluding 30)', '30 - 40 (excluding 40)', '40 - 50 (excluding 50)',
                              '50 - 60 (excluding 60)', '60 - 70 (excluding 70)', '70 - 80 (excluding 80)',
                              '80 - 90 (excluding 90)', '90 - 100']

        print_wm_header_dict = {'WM': 'wm', 'Class Rank': 'rank', '及格科目': 'rank'}

        pdf_merger_master = PdfFileMerger()
        pdf_merger_ct = PdfFileMerger()

        analysis_file = '2021-final-s456-analysis_Final.xlsm'
        exam_type = analysis_file.split('-')[1]
        analysis_open = os.path.join(self.db_folder, analysis_file)

        student_df = pd.read_excel(analysis_open, sheet_name='print')
        statistics_df = pd.read_excel(analysis_open, sheet_name='stat')

        for classlevel, classlist in senior_class_dict.items():
            passmark = 50 if classlevel == "s4" else 40
            top = 10
            for classcode in classlist:

                class_df = student_df[student_df['classcode'] == classcode].copy()
                statistics_col = ['item'] + mark_col_ut_dict[classlevel] + ['wm']
                stat_temp_df = statistics_df[statistics_df['classcode'] == classcode][statistics_col]
                class_stat_dict = stat_temp_df.to_dict('records')

                # stu_g   chi_g	eng_g	mth_g	chs_g	hst_g   geo_g	eco_g	phy_g	chm_g	bio_g   wm_g
                print_stuinfo = ['stuinfo_g']
                class_df['stuinfo_g'] = [[x1, x2, x3] for x1, x2, x3 in zip(class_df['key'],
                                                                            class_df['chname'],
                                                                            class_df['sex'])]
                # need class student counter
                print_stu_header = ['Class/No.', 'Chi Name', 'Sex']
                print_stu_header_dict = {'Class/No.': 'key', 'Chi Name': 'chname', 'Sex': 'sex'}
                print_subject_header_dict = {}
                print_subject = [col + '_g' for col in mark_col_ut_dict[classlevel]]

                # build subject item group
                for subject in mark_col_ut_dict[classlevel]:
                    rank_col = subject.title() + ' Form Rank'
                    # rank_col = 'Form Rank'
                    # add a selector for bik grade later
                    if subject in mark_col_ut_dict[classlevel]:
                        # print(subject)
                        print_subject_header_dict[subject.title()] = '"mark"'
                        print_subject_header_dict[rank_col] = '"rank"'
                        # build subject item group
                        class_df[subject+'_s'] = class_df[subject].apply(lambda x: '{:.2f}'.format(x))
                        class_df[subject + '_take'] = class_df[subject].isnull()

                        class_df[subject+'_p'] = class_df[subject].apply(
                            lambda x: '"mark"' if x >= passmark else '"mark fail"')
                        class_df[rank_col+'_s'] = (pd.to_numeric(class_df[rank_col], errors='coerce')
                                                     .fillna(-1).astype('int'))
                        class_df[rank_col+'_p'] = class_df[rank_col].apply(
                            lambda x: '"rank top"' if x <= top else '"rank"')
                        # build subject group
                        class_df[subject + '_g'] = [[x1, x2, x3, x4, x5] for x1, x2, x3, x4, x5 in
                                                    zip(class_df[subject+'_take'], class_df[subject+'_s'],
                                                        class_df[subject+'_p'],
                                                        class_df[rank_col+'_s'], class_df[rank_col+'_p'])]
                    else:
                        pass

                # build wm item group
                class_df['wm_p'] = class_df['wm'].apply(lambda x: '"mark"' if x >= passmark else '"mark fail"')
                class_df['crank_p'] = class_df['Class Rank'].apply(lambda x: '"rank top"' if x <= 10 else '"rank"')
                # class_df['frank_p'] = class_df['Form Rank'].apply(lambda x: '"rank top"' if x <= top else '"rank"')
                class_df['wm_g'] = [[x1, x2, x3, x4, x5] for x1, x2, x3, x4, x5 in zip(
                    class_df['wm'].apply(lambda x: '{:.2f}'.format(x)),
                    class_df['wm_p'],
                    class_df['Class Rank'].astype(int),
                    class_df['crank_p'],
                    # class_df['Form Rank'].astype(int),
                    # class_df['frank_p'],
                    class_df['及格科目'].astype(int))]

                print_wm = ['wm_g']
                print_col = print_stuinfo + print_subject + print_wm
                # class_df.to_excel('test.xlsx', index=False)
                # class_df[print_col].to_excel('print.xlsx', index=False)

                student_dict = class_df[print_col].to_dict('records')

                # setup for jinja2
                html_analysis_folder = os.path.join(self.html_template_root_folder, 'analysis')
                env = Environment(loader=FileSystemLoader(html_analysis_folder))
                template = env.get_template('analysis_template_senior.html')
                css = os.path.join(html_analysis_folder, 'analysis_template_senior.css')

                print_date = 'Print: {} ({})'.format(datetime.datetime.now().strftime('%Y-%m-%d'),
                                                     datetime.datetime.now().strftime('%H:%M'))
                class_info_dict = self.ct_dict[classcode.lower()]
                class_info_dict['exam_year'] = self.exam_year
                class_info_dict['exam_name'] = exam_type.title()
                class_info_dict['classcode'] = classcode.upper()
                class_info_dict['print_date'] = print_date

                template_vars = {'students': student_dict,
                                 'class_dict': class_info_dict,
                                 'statistics': class_stat_dict,
                                 'statistics_header': statistics_col,
                                 'statistics_col': stat_col_type_list,
                                 'print_stu_header': print_stu_header_dict,
                                 'print_subject_header': print_subject_header_dict,
                                 'print_wm_header': print_wm_header_dict,
                                 'print_subjects': print_subject,
                                 'print_stuinfo': print_stuinfo,
                                 'print_wm': print_wm,
                                 'css': css,
                                 }

                # print(print_stu_header_dict)
                # print(print_subject_header_dict)
                # print(print_wm_header_dict)

                options = {'page-size': 'A4',
                           'orientation': 'landscape',
                           'margin-top': '1.0cm',
                           'margin-right': '0.8cm',
                           'margin-bottom': '0.5cm',
                           'margin-left': '0.8cm',
                           'encoding': "UTF-8",
                           # 'encoding': 'ISO-8859-1',
                           'no-outline': None,
                           'enable-local-file-access': None}

                html_out = template.render(template_vars)
                html_save = os.path.join(self.db_folder, 'dump', 'temp.html')
                pdf_save = os.path.join(self.db_folder, 'dump',
                                        self.exam_year + '_' + exam_type + '_'
                                        + self.assessment + '_' + classcode.lower() + '.pdf')

                with open(html_save, 'w', encoding='utf-8') as f:
                    f.write(html_out)
                    f.close()

                pdfkit.from_file(html_save, pdf_save, css=css, options=options)
                print('{}: {} saved.'.format(classcode, pdf_save))

                pdf_merger_master.append(pdf_save)
                pdf_merger_ct.append(pdf_save)
                pdf_merger_ct.append(pdf_save)

        analysis_file_master = os.path.join(self.assessment_folder, 'pdf',
                                            self.exam_year + '_' + exam_type + '_' + 'analysis_senior.pdf')
        analysis_file_ct = os.path.join(self.assessment_folder, 'pdf',
                                        self.exam_year + '_' + exam_type + '_' + 'analysis_senior_ct.pdf')

        with open(analysis_file_master, 'wb') as file_object:
            pdf_merger_master.write(file_object)

        with open(analysis_file_ct, 'wb') as file_object:
            pdf_merger_ct.write(file_object)

        print(f'class analysis pdf generated: { analysis_file_master }')

    def trash(self):
        work_col_ext = ['key', 'regno', 'enname', 'chname', 'sex', 'classlevel', 'classcode', 'classno',
                        'chi', 'eng', 'mth', 'chs', 'hst', 'geo', 'eco', 'isc', 'phy', 'chm', 'bio', 'bik', 'lst',
                        'pth', 'cps', 'via', 'mus', 'ped', 'dte', 'hec', 'x',
                        'wm', 'wm1', 'improved', '1200', 'improved2', 'temp',
                        'Chi Form Rank', 'Eng Form Rank', 'Mth Form Rank', 'Chs Form Rank', 'Hst Form Rank',
                        'Geo Form Rank', 'Eco Form Rank', 'Isc Form Rank', 'Phy Form Rank', 'Chm Form Rank',
                        'Bio Form Rank', 'Bik Form Rank', 'Lst Form Rank', 'Pth Form Rank', 'Cps Form Rank',
                        'Via Form Rank', 'Mus Form Rank', 'Ped Form Rank', 'Dte Form Rank', 'Hec Form Rank',
                        'x rank', 'Class Rank', 'Form Rank', '及格科目', 'Improved Rank',
                        'spare1', 'spare2', 'spare3', 'spare4', 'spare5', 'spare6',
                        'chi_r', 'eng_r', 'mth_r', 'chs_r', 'hst_r', 'geo_r', 'eco_r', 'isc_r', 'phy_r', 'chm_r',
                        'bio_r', 'bik_r', 'lst_r', 'pth_r', 'cps_r', 'via_r', 'mus_r', 'ped_r', 'dte_r', 'hec_r',
                        'x_r', 'wm_r', 'Filter', 'Core Subj Ind', 'Chi', 'Eng', 'Mth', 'H', 'No of Fail']

        work_col_full = ['key', 'regno', 'enname', 'chname', 'sex', 'classlevel', 'classcode', 'classno',
                         'chi', 'eng', 'mth', 'chs', 'hst',
                         'geo', 'eco', 'isc', 'phy', 'chm',
                         'bio', 'bik', 'lst', 'pth', 'cps',
                         'via', 'mus', 'ped', 'dte', 'hec',
                         'wm', 'wm1', 'improved',
                         'Chi Form Rank', 'Eng Form Rank', 'Mth Form Rank', 'Chs Form Rank', 'Hst Form Rank',
                         'Geo Form Rank', 'Eco Form Rank', 'Isc Form Rank', 'Phy Form Rank', 'Chm Form Rank',
                         'Bio Form Rank', 'Bik Form Rank', 'Lst Form Rank', 'Pth Form Rank', 'Cps Form Rank',
                         'Via Form Rank', 'Mus Form Rank', 'Ped Form Rank', 'Dte Form Rank', 'Hec Form Rank',
                         'Class Rank', 'Form Rank', '及格科目', 'Improved Rank']


