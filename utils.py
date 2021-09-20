import openpyxl
import os
import io
from PyPDF2 import PdfFileWriter
from PyPDF2 import PdfFileReader
from pdfminer3.layout import LAParams
from pdfminer3.pdfpage import PDFPage
from pdfminer3.pdfinterp import PDFResourceManager
from pdfminer3.pdfinterp import PDFPageInterpreter
from pdfminer3.converter import TextConverter

# external folder
def convert_pdf_to_txt(path, file_password):
    rsrcmgr = PDFResourceManager()
    retstr = io.StringIO()
    codec = 'utf-8'
    laparams = LAParams()
    device = TextConverter(rsrcmgr, retstr, codec=codec, laparams=laparams)
    fp = io.open(path, 'rb')
    interpreter = PDFPageInterpreter(rsrcmgr, device)
    password = file_password
    maxpages = 0
    caching = True
    pagenos=set()

    for page in PDFPage.get_pages(fp, pagenos, maxpages=maxpages,
                                  password=password, caching=caching, check_extractable=True):
        interpreter.process_page(page)

    text = retstr.getvalue()

    fp.close()
    device.close()
    retstr.close()
    return text


_EXAM_YEAR_ = '2021'
_EXAM_TYPE_ = '_Exam1'

examyear = _EXAM_YEAR_
examtype = _EXAM_TYPE_
ext = '.xlsx'

# external folder
report_folder =
external_folder = '/Volumes/BTHACK/_exam_file_/pdf/'
sep = os.sep
homefolder = os.path.abspath(os.path.join(os.getcwd(), '..')) + sep
examfolder = homefolder + examyear + 'exam' + sep
templatefolder = homefolder + 'template' + sep
setupfolder = homefolder + 'setup' + sep
# schrpt_import_folder = homefolder + 'schrpt' + sep + 'pdf' + sep + examtype + sep
schrpt_import_folder = external_folder + sep
schrpt_output_folder = external_folder + sep + 'pw' + sep
schrpt_output2_folder = external_folder + sep + 'nopw' + sep
schrpt_temp_folder = external_folder + sep + 'pw' + sep + 'temp' + sep

# file setup
pwfile = setupfolder + '1920stu.xlsx'

schrptexam = 'CDGFSS_School_Report_' + examyear + examtype
# load password file
wb = openpyxl.load_workbook(filename=pwfile, data_only=True)

# print(AllStudents)
# rpt_filename = {'s123': '1920-schrpt-s123-final.pdf',
#                's45': '1920-schrpt-s45-final.pdf'}
#rpt_filename = {'s45': '1920-schrpt-s45-final.pdf'}

rpt_filename = {'s123': '2021-schrpt-s123-final.pdf', 's45': '2021-schrpt-s45-exam1.pdf'}

pdf_sheet_header = ['key', 'regno', 'enname', 'chname', 'classlevel',
                  'classcode', 'classno', 'status', 'hkid', 'password', 'status', 'print']

pdf_ind = dict(zip(pdf_sheet_header, range(1,len(pdf_sheet_header)+1)))

def check():
    for key, item in rpt_filename.items():
        studb = wb[key]
        schrpt_src_file = schrpt_import_folder + item
        print('Start checking {}'.format(schrpt_src_file))
        SchRptOpen = PdfFileReader(schrpt_src_file)

        num_student = studb.max_row - 1
        num_pagesPdf = SchRptOpen.getNumPages()
        print('Checking (no. of student, no. of pages):', num_student, num_pagesPdf)

        if num_student != num_pagesPdf:
            print('PDF pages do not match the number of students.')
            exit(3)

        else:
            StuRange = range(2, studb.max_row + 1)
            resource_manager = PDFResourceManager()
            fake_file_handle = io.StringIO()
            converter = TextConverter(resource_manager, fake_file_handle, laparams=LAParams())
            page_interpreter = PDFPageInterpreter(resource_manager, converter)

            for row in StuRange:
                pagePdf = SchRptOpen.getPage(row - 2)
                stuname1 = studb.cell(row=row, column=stufile_ind['name1']).value
                stuname2 = studb.cell(row=row, column=stufile_ind['name2']).value
                classcode_no = studb.cell(row=row, column=stufile_ind['key']).value
                password = studb.cell(row=row, column=stufile_ind['searchid']).value
                pdf_writer = PdfFileWriter()
                pdf_writer.addPage(pagePdf)
                #report_temp = schrpt_temp_folder + sep + 'temp.pdf'
                report_temp = schrpt_output2_folder + sep + key + sep \
                                    + '{}_{}_{}.pdf'.format(schrptexam, classcode_no, stuname2)

                with io.open(report_temp, 'wb') as out:
                    pdf_writer.write(out)

                txt = convert_pdf_to_txt(report_temp, password)
                find_name = txt.find(stuname1)

                if find_name == -1:
                    print(stuname1, find_name)

                else:
                    student_name_in_pdf = txt[find_name:(len(stuname1) + find_name)]
                    print('Checked: #{} {} - {}'.format(row - 1, stuname1, find_name))


check()

for key, item in rpt_filename.items():
    studb = wb[key]
    schrpt_src_file = schrpt_import_folder + item
    print('Start splitting {}'.format(schrpt_src_file))
    SchRptOpen = PdfFileReader(schrpt_src_file)

    num_student = studb.max_row - 1
    num_pagesPdf = SchRptOpen.getNumPages()
    print(num_student, num_pagesPdf)

    if num_student != num_pagesPdf:
        print('PDF pages do not match the number of students.')
        exit(3)

    else:
        StuRange = range(2, studb.max_row + 1)
        temp_range = range(2, 10)

        for row in StuRange:
            pagePdf = SchRptOpen.getPage(row-2)
            stuname1 = studb.cell(row=row, column=stufile_ind['name1']).value
            stuname2 = studb.cell(row=row, column=stufile_ind['name2']).value
            classcode_no = studb.cell(row=row, column=stufile_ind['key']).value
            password = studb.cell(row=row, column=stufile_ind['searchid']).value
            pdf_writer = PdfFileWriter()
            pdf_writer.addPage(pagePdf)
            report_file_split = schrpt_output_folder + sep + key + sep \
                + '{}_{}_{}.pdf'.format(schrptexam, classcode_no, stuname2)

            with io.open(report_file_split, 'wb') as out:
                pdf_writer.encrypt(password)
                pdf_writer.write(out)

            # txt = convert_pdf_to_txt(report_file_split, password)
            # find_name = txt.find(stuname1)

            # if find_name == -1:
            #     print(report_file_split + ' have some error. Must check.')

            # else:
            #    student_name_in_pdf = txt[find_name:(len(stuname1) + find_name)]
            print('Created: #{} {}'.format(row - 1, report_file_split))

            # print('Created: {}'.format(report_file_split))


